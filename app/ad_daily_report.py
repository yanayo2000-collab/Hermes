from __future__ import annotations

import hashlib
import io
import json
import random
import re
import sqlite3
from dataclasses import asdict, dataclass, field, replace
from datetime import date, datetime, time, timedelta, timezone
from typing import Any, Dict, Iterable, List, Optional, Protocol, Tuple
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.tugao_bi import query_tugao_bind_success_rows
from app.ad_creative_intelligence import build_creative_intelligence_payload
from app.ad_recommendation_v4 import RULE_VERSION as RECOMMENDATION_RULE_VERSION, score_ad_object


REPORT_SCHEMA_VERSION = 'ad_daily_report_v1'
AD_OBSERVATION_IDENTITY_VERSION = 'ad_observation_location_v1'
HISTORICAL_SETTLEMENT_ACCOUNT_LABELS = {
    'archived_settled_accounts',
    '历史失权账户归档（已结算）',
}
HISTORICAL_SETTLEMENT_PLACEHOLDER = '历史已结算事实保全'
TUGAO_REAL_BIND_DEDUPE_VERSION = 'tugao_success_unique_customer_user_v1'
TUGAO_REAL_BIND_ATTRIBUTION_VERSION = 'tugao_ad_grain_exact_v1'
REAL_BIND_COUNT_MODE = 'unique_customer_user_or_bind'
REAL_BIND_COUNT_MODE_LABEL_CN = '唯一业务用户优先，缺失时按唯一绑定计数'
REPORT_TZ = timezone.utc
REPORT_TZ_NAME = 'UTC'
UNRESOLVED_COUNTRY_LABEL = '待归属'


ZH_LABELS: Dict[str, str] = {
    'campaign': '广告系列',
    'ad_set': '广告组',
    'ad': '广告',
    'spend': '消耗金额',
    'impressions': '展示次数',
    'clicks': '点击次数',
    'ctr': '点击率',
    'cpm': '千次展示成本',
    'cpi': '安装成本',
    'real_bind_cpa': '真实入会成本',
    'winner': '正式赢家',
    'potential_winner': '潜力广告',
    'recommendation': '优化建议',
    'confidence': '建议置信度',
    'mixed_change': '混合调整',
    'under_delivery': '投放不足',
    'scale_up': '放量',
    'reduce_budget': '降预算',
    'pause': '暂停',
    'observe': '继续观察',
    'hold_scale': '暂缓放量',
    'manual_review': '人工复核',
    'sample_insufficient': '样本不足',
    'data_insufficient': '核心数据未接齐',
    'data_missing': '数据异常',
    'zero_stop': '零入会止损',
    'severe_over_cap': '严重超阈值',
    'over_cap': '超阈值',
    'slight_over_cap': '轻微超阈值',
    'no_cap': '无红线',
    'data_quality': '数据质量',
    'data_anomaly': '数据异常',
    'post_funnel_event_inconsistent': '后链路事件异常',
    'view_capability_missing': 'IM/后链路数据未接入',
    'sample_insufficient': '样本不足',
    'scale_opportunity': '可放量',
    'front_funnel_weak': '素材前链路弱',
    'creative_fatigue': '素材疲劳',
    'low_quality_traffic': '点击诱导 / 低质量流量',
    'audience_mismatch': '人群不匹配',
    'creative_effective_post_im_failed': '素材有效但后链路失败',
    'business_result_anomaly': '经营结果异常',
    'im_handoff_issue': 'IM 承接问题',
    'cs_response_issue': '客服响应问题',
    'linky_crm_issue': 'Linky / bind / CRM 问题',
    'creative_scale_candidate': '素材放量候选',
    'continue_observe': '继续观察',
    'generate_repair_creative': '生成修正素材',
    'generate_derivative_creative': '生成衍生素材',
    'inspect_post_im_funnel': '检查im链路',
    'check_im_flow': '检查im链路',
    'check_linky_bind_crm': '检查 Linky / bind / CRM',
    'inspect_business_result': '检查经营结果',
    'inspect_data_quality': '检查数据质量',
    'check_data_mapping': '检查数据映射',
    'check_timo_im_mapping': '检查 Timo IM 映射',
    'check_linky_bind_crm_tracking': '检查 Linky/bind 事件',
}


LEGACY_STATUS_DIAGNOSIS_TYPE: Dict[str, str] = {
    'data_missing': 'data_anomaly',
    'data_quality': 'data_anomaly',
    'winner': 'scale_opportunity',
    'potential_winner': 'continue_observe',
    'frontend_risk': 'creative_fatigue',
    'zero_stop': 'front_funnel_weak',
    'severe_over_cap': 'front_funnel_weak',
    'over_cap': 'front_funnel_weak',
    'slight_over_cap': 'continue_observe',
    'sample_insufficient': 'sample_insufficient',
    'under_delivery': 'sample_insufficient',
    'no_cap': 'continue_observe',
    'mixed_change': 'continue_observe',
    'structure_optimization': 'front_funnel_weak',
}

LEGACY_GENERATIVE_STATUS_TAGS = {
    'winner',
    'potential_winner',
    'frontend_risk',
    'over_cap',
    'severe_over_cap',
    'zero_stop',
    'structure_optimization',
}


def _diagnosis_type_from_status(status_tag: str) -> str:
    return LEGACY_STATUS_DIAGNOSIS_TYPE.get(str(status_tag or '').strip(), 'continue_observe')


def _action_type_from_diagnosis(primary_action: str, diagnosis_type: str) -> str:
    diagnosis = str(diagnosis_type or '').strip()
    if diagnosis == 'scale_opportunity':
        return 'generate_derivative_creative'
    if diagnosis in {'front_funnel_weak', 'low_quality_traffic', 'creative_fatigue', 'audience_mismatch'}:
        return 'generate_repair_creative'
    if diagnosis in {'creative_effective_post_im_failed', 'im_handoff_issue', 'cs_response_issue', 'linky_crm_issue'}:
        return 'inspect_post_im_funnel'
    if diagnosis == 'business_result_anomaly':
        return 'inspect_business_result'
    if diagnosis == 'data_anomaly':
        return 'inspect_data_quality'
    return str(primary_action or 'observe')


DEFAULT_RULE_CONFIG: Dict[str, Any] = {
    'markets': {
        'ID': {'real_bind_cpa_cap': 0.70},
        'BR': {'real_bind_cpa_cap': 1.20},
        'RECOMPA': {'real_bind_cpa_cap': None},
    },
    'rules': {
        'min_spend_for_strong_action': 5.0,
        'min_formal_winner_binds': 10,
        'high_confidence_binds': 20,
        'zero_conversion_spend_multiplier': 3,
        'sample_insufficient_spend_multiplier': 1,
        'maturity_day': 2,
        'under_delivery_review_day': 4,
        'budget_change_cooldown_hours': 72,
        'frontend_risk': {
            'cpi_increase_pct': 25,
            'ctr_decrease_pct': 20,
            'cpm_increase_pct': 30,
            'min_trigger_count': 2,
        },
        'sample_thresholds': {
            'default': {
                'min_spend': 0.0,
                'min_im': 20,
                'min_apply_for_rate_diagnosis': 50,
                'min_user_engaged_im': 10,
                'min_high_intent_im': 5,
                'min_join': 5,
                'min_observation_hours': 24,
            },
        },
        'funnel_baselines': {
            'BR': {
                'im_cost_p50': 0.43,
                'registration_to_apply_rate_normal': 0.70,
                'user_engaged_im_cost_p50': 1.6,
                'high_intent_im_cost_p50': 2.8,
                'high_value_rate_p50': 0.60,
                'user_engaged_im_rate_p50': 0.24,
                'high_intent_im_rate_p50': 0.35,
                'im_to_join_rate_p50': 0.09,
                'im_to_join_rate_warning': 0.08,
            },
            'ID': {
                'im_cost_p50': 0.42,
                'registration_to_apply_rate_normal': 0.60,
                'user_engaged_im_cost_p50': 1.6,
                'high_intent_im_cost_p50': 2.2,
                'high_value_rate_p50': 0.45,
                'user_engaged_im_rate_p50': 0.25,
                'high_intent_im_rate_p50': 0.30,
                'im_to_join_rate_p50': 0.19,
                'im_to_join_rate_warning': 0.16,
            },
        },
    },
}


COUNTRY_CODE_ALIASES = {
    'indonesia': 'ID',
    'id': 'ID',
    'idn': 'ID',
    '印尼': 'ID',
    'brazil': 'BR',
    'brasil': 'BR',
    'br': 'BR',
    '巴西': 'BR',
    'mexico': 'MX',
    'méxico': 'MX',
    'mx': 'MX',
    '墨西哥': 'MX',
    'colombia': 'CO',
    'co': 'CO',
    '哥伦比亚': 'CO',
    'venezuela': 'VE',
    've': 'VE',
    '委内瑞拉': 'VE',
    'recompa': 'RECOMPA',
}

COUNTRY_HINT_PATTERNS: Tuple[Tuple[str, Tuple[str, ...]], ...] = (
    ('BR', ('brazil', 'brasil', '巴西', 'br01', 'br-')),
    ('ID', ('indonesia', 'indo', '印尼', 'id01', 'idn', 'id-')),
    ('MX', ('mexico', 'méxico', '墨西哥', 'mx01', 'mx-')),
    ('CO', ('colombia', '哥伦比亚', 'co01', 'co-')),
    ('VE', ('venezuela', '委内瑞拉', 've01', 've-')),
)


@dataclass(frozen=True)
class DataQualityStatus:
    status: str = 'ok'
    freshness: str = 'fresh'
    attribution_quality: str = 'fixture'
    warnings: List[str] = field(default_factory=list)


@dataclass(frozen=True)
class RealBindMetricContract:
    dedupe_version: str
    attribution_version: str
    real_bind_count_mode: str
    real_bind_count_mode_label_cn: str
    is_dedupe_confirmed: bool
    bind_event_count: int
    unique_bind_count: int
    unique_customer_user_count: int
    final_real_bind_count: int
    has_wa_success_count: int
    no_wa_success_count: int
    duplicate_event_id_count: int = 0
    duplicate_bind_id_count: int = 0
    duplicate_customer_user_id_count: int = 0
    missing_country_count: int = 0
    missing_project_count: int = 0
    missing_attribution_count: int = 0


@dataclass(frozen=True)
class DataQualityGateResult:
    status: str
    status_zh: str
    reasons: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    checked_at_utc: str = ''


@dataclass(frozen=True)
class BudgetMode:
    mode: str = 'unknown'
    last_budget_change_at: Optional[str] = None
    cooldown_hours_remaining: int = 0


@dataclass(frozen=True)
class AdObjectMetrics:
    object_id: str
    object_level: str
    country: str
    project: str
    account_id: str
    campaign: str
    ad_group: str
    ad: str
    spend: float
    impressions: float
    clicks: float
    ctr: float
    cpm: float
    installs: float
    cpi: Optional[float]
    target_app: str = 'inactive'
    registrations: float = 0.0
    high_value_users: float = 0.0
    im_entries: float = 0.0
    auto_apply_user_count: float = 0.0
    auto_apply_message_users: float = 0.0
    auto_apply_user_count_source: str = ''
    user_first_reply_users: float = 0.0
    im_user_message_ge_3_users: float = 0.0
    im_user_message_ge_5_users: float = 0.0
    link_click_users: float = 0.0
    linky_register_users: float = 0.0
    bind_success_users: float = 0.0
    crm_succeed_users: float = 0.0
    system_touched_im_users: float = 0.0
    user_engaged_im_users: float = 0.0
    high_intent_im_users: float = 0.0
    high_value_rate: Optional[float] = None
    registration_to_apply_rate: Optional[float] = None
    im_cost: Optional[float] = None
    system_touched_im_rate: Optional[float] = None
    user_engaged_im_rate: Optional[float] = None
    user_engaged_im_cost: Optional[float] = None
    user_engaged_im_metric_version: str = 'unknown'
    high_intent_im_rate: Optional[float] = None
    high_intent_im_cost: Optional[float] = None
    im_to_join_rate: Optional[float] = None
    linky_register_rate_from_link_click: Optional[float] = None
    bind_rate_from_linky: Optional[float] = None
    crm_succeed_rate_from_bind: Optional[float] = None
    af_guild_joins: float = 0.0
    real_bind_count: int = 0
    real_bind_cpa: Optional[float] = None
    first_effective_spend_date: str = ''
    maturity_day: int = 0
    data_quality: DataQualityStatus = field(default_factory=DataQualityStatus)
    budget_mode: BudgetMode = field(default_factory=BudgetMode)
    frontend_trend: Dict[str, float] = field(default_factory=dict)
    metric_availability: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    source_type: str = ''
    observation_identity: str = ''


@dataclass(frozen=True)
class RealBindEvent:
    event_id: str
    occurred_at_utc: str
    country: str
    project: str
    account_id: str
    campaign: str
    ad_group: str
    ad: str
    user_key: str
    attribution_quality: str = 'fixture'
    is_duplicate: bool = False
    bind_status: str = 'success'
    has_wa: Optional[bool] = None
    bind_id: str = ''
    customer_user_id: str = ''
    source_updated_at: str = ''


@dataclass(frozen=True)
class RecommendationEvidence:
    spend: float
    real_bind_count: int
    real_bind_cpa: Optional[float]
    cpi: Optional[float]
    ctr: float
    cpm: float
    country_cap: Optional[float]
    data_window: Dict[str, str]
    data_quality: DataQualityStatus
    budget_mode: BudgetMode
    frontend_trend: Dict[str, float]
    installs: float = 0.0
    funnel_metrics: Dict[str, Any] = field(default_factory=dict)
    evidence_points: List[str] = field(default_factory=list)
    scorecard: Dict[str, Any] = field(default_factory=dict)
    rule_version: str = RECOMMENDATION_RULE_VERSION


@dataclass(frozen=True)
class Recommendation:
    recommendation_id: str
    object_id: str
    object_level: str
    object_name: str
    country: str
    project: str
    primary_action: str
    primary_action_zh: str
    adjustment_pct: int
    reason_zh: str
    confidence: str
    confidence_zh: str
    status_tag: str
    evidence: RecommendationEvidence
    created_at_utc: str
    diagnosis_type: str = 'continue_observe'
    diagnosis_type_zh: str = '继续观察'
    action_type: str = 'observe'
    action_type_zh: str = '继续观察'
    primary_layer: str = 'unknown'
    maturity_status: str = 'unknown'
    data_quality_status: str = 'ok'
    allow_pause: bool = False
    allow_scale: bool = False
    creative_scale_candidate: bool = False
    business_scale_allowed: bool = False
    allow_generate_creative: bool = False
    needs_data: List[str] = field(default_factory=list)
    creative_diagnosis: Dict[str, Any] = field(default_factory=dict)
    post_im_diagnosis: Dict[str, Any] = field(default_factory=dict)
    business_diagnosis: Dict[str, Any] = field(default_factory=dict)
    action_gate: Dict[str, Any] = field(default_factory=dict)
    decision_context: Dict[str, Any] = field(default_factory=dict)
    data_origin: str = 'NATIVE_V2'
    observation_identity: str = ''


@dataclass(frozen=True)
class ProductionRecommendationGateResult:
    recommendation_id: str
    gate_status: str
    publishable: bool
    mode: str
    reasons: List[str] = field(default_factory=list)
    checked_at_utc: str = ''


@dataclass(frozen=True)
class DailyAdReportV1:
    report_id: str
    snapshot_version: str
    rule_version: str
    report_date: str
    window_start_utc: str
    window_end_utc: str
    window_timezone: str
    data_mode: str
    provider: str
    generated_at_utc: str
    summary: Dict[str, Any]
    real_bind_metric: RealBindMetricContract
    data_quality_gate: DataQualityGateResult
    ad_objects: List[AdObjectMetrics]
    recommendations: List[Recommendation]
    creative_test_plan: List[Dict[str, Any]]
    creative_insights: Dict[str, Any]
    review_skeleton: List[Dict[str, Any]]
    labels: Dict[str, str] = field(default_factory=lambda: dict(ZH_LABELS))
    simulation_notice: str = '当前为模拟真实绑定数据，仅用于系统验证，不代表生产投放结论。'


class RealConversionProvider(Protocol):
    def get_bind_events(
        self,
        start_time: datetime,
        end_time: datetime,
        project: Optional[str] = None,
        country: Optional[str] = None,
    ) -> List[RealBindEvent]:
        ...


def evaluate_production_recommendation_gate(
    recommendation: Recommendation,
    *,
    mode: str = 'shadow',
    allowed_actions: Optional[Iterable[str]] = None,
    data_quality_gate: Optional[DataQualityGateResult] = None,
    require_dedupe_confirmed: bool = True,
) -> ProductionRecommendationGateResult:
    normalized_mode = str(mode or 'shadow').strip().lower() or 'shadow'
    allowed = {str(item or '').strip() for item in (allowed_actions or ['scale_up', 'reduce_budget', 'pause']) if str(item or '').strip()}
    reasons: List[str] = []

    if normalized_mode != 'production':
        reasons.append('shadow_mode')
    if recommendation.primary_action not in allowed:
        reasons.append('action_not_publishable')
    if data_quality_gate is not None and data_quality_gate.status != 'PASS':
        reasons.append('data_quality_gate_not_pass')
    data_quality = recommendation.evidence.data_quality
    if data_quality.status != 'ok':
        reasons.append('data_quality_not_ok')
    if data_quality.attribution_quality in {'fixture', 'simulated', 'unknown', ''}:
        reasons.append('attribution_not_verified')
    if require_dedupe_confirmed and data_quality_gate is not None and 'dedupe_not_confirmed' in set(data_quality_gate.warnings):
        reasons.append('dedupe_not_confirmed')
    budget_mode = recommendation.evidence.budget_mode
    if budget_mode.mode == 'mixed_change' or int(budget_mode.cooldown_hours_remaining or 0) > 0:
        reasons.append('budget_cooldown_or_mixed_change')
    if recommendation.confidence == 'low':
        reasons.append('low_confidence')

    publishable = not reasons
    return ProductionRecommendationGateResult(
        recommendation_id=recommendation.recommendation_id,
        gate_status='publishable' if publishable else 'blocked',
        publishable=publishable,
        mode=normalized_mode,
        reasons=reasons,
        checked_at_utc=datetime.now(timezone.utc).isoformat(),
    )


def london_report_window(report_date: Optional[str] = None) -> Tuple[date, datetime, datetime]:
    if report_date:
        target = datetime.strptime(str(report_date), '%Y-%m-%d').date()
    else:
        target = datetime.now(REPORT_TZ).date()
    start_local = datetime.combine(target, time.min, tzinfo=REPORT_TZ)
    end_local = start_local + timedelta(days=1)
    return target, start_local.astimezone(timezone.utc), end_local.astimezone(timezone.utc)


def normalize_country_code(value: Any) -> str:
    text = str(value or '').strip()
    if not text:
        return 'UNKNOWN'
    if text == UNRESOLVED_COUNTRY_LABEL:
        return UNRESOLVED_COUNTRY_LABEL
    lowered = text.lower()
    if lowered in COUNTRY_CODE_ALIASES:
        return COUNTRY_CODE_ALIASES[lowered]
    if 'indonesia' in lowered or '印尼' in lowered:
        return 'ID'
    if 'brazil' in lowered or 'brasil' in lowered or '巴西' in lowered:
        return 'BR'
    if 'recompa' in lowered:
        return 'RECOMPA'
    return text.upper()


def country_is_unresolved(value: Any) -> bool:
    return normalize_country_code(value) in {'UNKNOWN', UNRESOLVED_COUNTRY_LABEL}


def infer_country_from_ad_context(*values: Any) -> str:
    raw = ' '.join(str(value or '') for value in values if str(value or '').strip())
    if not raw:
        return ''
    normalized = raw.lower()
    for country, hints in COUNTRY_HINT_PATTERNS:
        for hint in hints:
            if hint in normalized:
                return country
    compact_tokens = set(re.findall(r'[a-z]{2,}', normalized))
    if 'br' in compact_tokens:
        return 'BR'
    if 'id' in compact_tokens:
        return 'ID'
    if 'mx' in compact_tokens:
        return 'MX'
    if 'co' in compact_tokens:
        return 'CO'
    if 've' in compact_tokens:
        return 'VE'
    return ''


def _stable_id(*parts: Any) -> str:
    raw = '|'.join(str(part or '') for part in parts)
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()[:16]


def ad_observation_identity(
    *,
    platform: Any,
    account_id: Any,
    country: Any,
    campaign: Any,
    ad_group: Any,
    ad: Any,
    source_type: Any,
) -> str:
    """Stable dashboard-row identity, independent of report and recommendation revisions."""
    return _stable_id(
        AD_OBSERVATION_IDENTITY_VERSION,
        str(platform or '').strip().casefold(),
        str(account_id or '').strip().casefold(),
        normalize_country_code(country).casefold(),
        str(campaign or '').strip().casefold(),
        str(ad_group or '').strip().casefold(),
        str(ad or '').strip().casefold(),
        str(source_type or '').strip().casefold(),
    )


class FixtureRealConversionProvider:
    """Deterministic shadow provider used until TUGAO real bind data is available."""

    FIXTURE_SCENARIOS: Tuple[Dict[str, Any], ...] = (
        {'scenario': 'id_formal_winner', 'country': 'ID', 'spend': 14.56, 'binds': 26, 'campaign': 'ID 正式赢家', 'ad_group': '核心组 A', 'ad': 'winner-26', 'first_day_offset': 4},
        {'scenario': 'id_edge_winner', 'country': 'ID', 'spend': 6.90, 'binds': 10, 'campaign': 'ID 临界赢家', 'ad_group': '核心组 B', 'ad': 'winner-10', 'first_day_offset': 3},
        {'scenario': 'br_strong_winner', 'country': 'BR', 'spend': 14.40, 'binds': 18, 'campaign': 'BR 强赢家', 'ad_group': 'Brazil A', 'ad': 'winner-18', 'first_day_offset': 5},
        {'scenario': 'id_slight_over', 'country': 'ID', 'spend': 8.88, 'binds': 12, 'campaign': 'ID 轻微超标', 'ad_group': 'Review A', 'ad': 'slight-over', 'first_day_offset': 3},
        {'scenario': 'id_mid_over', 'country': 'ID', 'spend': 10.20, 'binds': 12, 'campaign': 'ID 中度超标', 'ad_group': 'Review B', 'ad': 'mid-over', 'first_day_offset': 3},
        {'scenario': 'id_severe_over', 'country': 'ID', 'spend': 10.00, 'binds': 10, 'campaign': 'ID 严重超标', 'ad_group': 'Risk A', 'ad': 'severe-over', 'first_day_offset': 3},
        {'scenario': 'id_zero_stop', 'country': 'ID', 'spend': 2.20, 'binds': 0, 'campaign': 'ID 零转化止损', 'ad_group': 'Risk B', 'ad': 'zero-stop', 'first_day_offset': 2},
        {'scenario': 'id_under_delivery', 'country': 'ID', 'spend': 0.80, 'binds': 0, 'campaign': 'ID 投放不足', 'ad_group': 'Learning A', 'ad': 'under-delivery', 'first_day_offset': 4},
        {'scenario': 'id_potential', 'country': 'ID', 'spend': 1.00, 'binds': 2, 'campaign': 'ID 潜力广告', 'ad_group': 'Test A', 'ad': 'potential-2', 'first_day_offset': 1},
        {'scenario': 'id_frontend_risk', 'country': 'ID', 'spend': 9.80, 'binds': 14, 'campaign': 'ID 前端恶化', 'ad_group': 'Risk C', 'ad': 'frontend-risk', 'first_day_offset': 3, 'cpi_increase_pct': 35, 'ctr_decrease_pct': 30},
        {'scenario': 'recompa_no_cap', 'country': 'RECOMPA', 'spend': 8.00, 'binds': 8, 'campaign': 'RECOMPA 无红线', 'ad_group': 'Relative A', 'ad': 'no-cap', 'first_day_offset': 3},
        {'scenario': 'structure_optimization', 'country': 'ID', 'spend': 12.00, 'binds': 8, 'campaign': 'ID 结构优化', 'ad_group': 'Mixed A', 'ad': 'mixed-campaign', 'first_day_offset': 4, 'has_inner_winner': True},
        {'scenario': 'mixed_change', 'country': 'ID', 'spend': 6.30, 'binds': 9, 'campaign': 'ID 混合调整', 'ad_group': 'Operator Change', 'ad': 'mixed-change', 'first_day_offset': 3, 'mixed_change': True},
    )

    def __init__(self, *, seed: int = 20260622, random_count: int = 0) -> None:
        self.seed = seed
        self.random_count = max(0, int(random_count or 0))

    def get_bind_events(
        self,
        start_time: datetime,
        end_time: datetime,
        project: Optional[str] = None,
        country: Optional[str] = None,
    ) -> List[RealBindEvent]:
        wanted_country = normalize_country_code(country) if country else ''
        events: List[RealBindEvent] = []
        for scenario in self.FIXTURE_SCENARIOS:
            if wanted_country and scenario['country'] != wanted_country:
                continue
            for index in range(int(scenario.get('binds') or 0)):
                user_key = f"{scenario['scenario']}-user-{index:03d}"
                events.append(RealBindEvent(
                    event_id=_stable_id(scenario['scenario'], user_key),
                    occurred_at_utc=(start_time + timedelta(minutes=index + 1)).isoformat(),
                    country=str(scenario['country']),
                    project=str(project or 'fixture_project'),
                    account_id='fixture_meta_account',
                    campaign=str(scenario['campaign']),
                    ad_group=str(scenario['ad_group']),
                    ad=str(scenario['ad']),
                    user_key=user_key,
                    attribution_quality='fixture_exact',
                    is_duplicate=False,
                ))
            if int(scenario.get('binds') or 0) > 0:
                events.append(RealBindEvent(
                    event_id=_stable_id(scenario['scenario'], 'duplicate'),
                    occurred_at_utc=(start_time + timedelta(minutes=999)).isoformat(),
                    country=str(scenario['country']),
                    project=str(project or 'fixture_project'),
                    account_id='fixture_meta_account',
                    campaign=str(scenario['campaign']),
                    ad_group=str(scenario['ad_group']),
                    ad=str(scenario['ad']),
                    user_key=f"{scenario['scenario']}-user-000",
                    attribution_quality='fixture_duplicate',
                    is_duplicate=True,
                ))
        rnd = random.Random(self.seed)
        for index in range(self.random_count):
            country_code = rnd.choice(['ID', 'BR'])
            if wanted_country and country_code != wanted_country:
                continue
            campaign = f'{country_code} 随机验证 {index + 1}'
            binds = rnd.randint(0, 24)
            for bind_index in range(binds):
                events.append(RealBindEvent(
                    event_id=_stable_id('random', index, bind_index),
                    occurred_at_utc=(start_time + timedelta(minutes=1200 + bind_index)).isoformat(),
                    country=country_code,
                    project=str(project or 'fixture_project'),
                    account_id='fixture_meta_account',
                    campaign=campaign,
                    ad_group='随机测试组',
                    ad=f'random-{index + 1}',
                    user_key=f'random-{index}-{bind_index}',
                    attribution_quality='fixture_random',
                ))
        return [event for event in events if start_time <= datetime.fromisoformat(event.occurred_at_utc) < end_time]


class TugaoRealConversionProvider:
    def __init__(
        self,
        *,
        db_path: Optional[str] = None,
        connection_factory: Optional[Any] = None,
    ) -> None:
        self.db_path = db_path
        self.connection_factory = connection_factory

    def _connect(self) -> sqlite3.Connection:
        if self.connection_factory is not None:
            return self.connection_factory()
        if not self.db_path:
            raise RuntimeError('tugao_provider_db_not_configured')
        conn = sqlite3.connect(self.db_path, timeout=30.0)
        conn.row_factory = sqlite3.Row
        return conn

    def get_bind_events(
        self,
        start_time: datetime,
        end_time: datetime,
        project: Optional[str] = None,
        country: Optional[str] = None,
    ) -> List[RealBindEvent]:
        conn = self._connect()
        should_close = self.connection_factory is None
        try:
            rows = query_tugao_bind_success_rows(
                conn,
                start_time=start_time.isoformat(),
                end_time=end_time.isoformat(),
                project=project,
                country=normalize_country_code(country) if country else None,
            )
            events: List[RealBindEvent] = []
            for row in rows:
                occurred_at = str(row['occurred_at_utc'] or row['updated_at_utc'] or row['business_date'] or '')
                events.append(RealBindEvent(
                    event_id=str(row['event_id'] or ''),
                    occurred_at_utc=occurred_at,
                    country=normalize_country_code(row['country']),
                    project=str(row['project'] or project or ''),
                    account_id=str(row['media_source'] or ''),
                    campaign=str(row['campaign_name'] or row['campaign_id'] or ''),
                    ad_group=str(row['adset_name'] or row['adset_id'] or ''),
                    ad=str(row['ad_name'] or row['ad_id'] or ''),
                    user_key=str(row['customer_user_id'] or row['bind_id'] or row['user_key'] or row['event_id'] or ''),
                    attribution_quality='tugao_raw_shadow',
                    is_duplicate=False,
                    bind_status=str(row['bind_status'] or 'success'),
                    has_wa=bool(row['has_wa']) if row['has_wa'] is not None else None,
                    bind_id=str(row['bind_id'] or ''),
                    customer_user_id=str(row['customer_user_id'] or ''),
                    source_updated_at=str(row['updated_at_utc'] or ''),
                ))
            return events
        finally:
            if should_close:
                conn.close()


def dedupe_first_effective_bind_events(events: Iterable[RealBindEvent]) -> List[RealBindEvent]:
    first_by_user: Dict[str, RealBindEvent] = {}
    for event in sorted(events, key=lambda item: item.occurred_at_utc):
        if event.is_duplicate:
            continue
        if str(event.bind_status or 'success').lower() != 'success':
            continue
        key = f'{event.country}|{event.project}|{event.user_key}'
        first_by_user.setdefault(key, event)
    return list(first_by_user.values())


def _duplicate_count(values: Iterable[str]) -> int:
    seen = set()
    duplicated = set()
    for value in values:
        text = str(value or '').strip()
        if not text:
            continue
        if text in seen:
            duplicated.add(text)
        seen.add(text)
    return len(duplicated)


def summarize_real_bind_metric_contract(
    events: Iterable[RealBindEvent],
    *,
    dedupe_confirmed: bool,
    dedupe_version: str = TUGAO_REAL_BIND_DEDUPE_VERSION,
    attribution_version: str = TUGAO_REAL_BIND_ATTRIBUTION_VERSION,
) -> RealBindMetricContract:
    success_events = [
        event for event in events
        if not event.is_duplicate and str(event.bind_status or 'success').lower() == 'success'
    ]
    bind_ids = [str(event.bind_id or '').strip() for event in success_events]
    customer_user_ids = [str(event.customer_user_id or event.user_key or '').strip() for event in success_events]
    unique_bind_ids = {value for value in bind_ids if value}
    unique_customer_user_ids = {value for value in customer_user_ids if value}
    final_count = len(unique_customer_user_ids) if unique_customer_user_ids else (len(unique_bind_ids) if unique_bind_ids else len(success_events))
    return RealBindMetricContract(
        dedupe_version=dedupe_version,
        attribution_version=attribution_version,
        real_bind_count_mode=REAL_BIND_COUNT_MODE,
        real_bind_count_mode_label_cn=REAL_BIND_COUNT_MODE_LABEL_CN,
        is_dedupe_confirmed=bool(dedupe_confirmed),
        bind_event_count=len(success_events),
        unique_bind_count=len(unique_bind_ids),
        unique_customer_user_count=len(unique_customer_user_ids),
        final_real_bind_count=final_count,
        has_wa_success_count=sum(1 for event in success_events if event.has_wa is True),
        no_wa_success_count=sum(1 for event in success_events if event.has_wa is False),
        duplicate_event_id_count=_duplicate_count(event.event_id for event in success_events),
        duplicate_bind_id_count=_duplicate_count(bind_ids),
        duplicate_customer_user_id_count=_duplicate_count(customer_user_ids),
        missing_country_count=sum(1 for event in success_events if normalize_country_code(event.country) == 'UNKNOWN'),
        missing_project_count=sum(1 for event in success_events if not str(event.project or '').strip()),
        missing_attribution_count=sum(1 for event in success_events if not (event.campaign or event.ad_group or event.ad)),
    )


def evaluate_data_quality_gate(
    metric: RealBindMetricContract,
    *,
    report_window_complete: bool = True,
    data_fresh: bool = True,
    sync_failed_count: int = 0,
    attribution_coverage: Optional[float] = None,
    min_attribution_coverage: float = 0.8,
) -> DataQualityGateResult:
    reasons: List[str] = []
    warnings: List[str] = []
    if not report_window_complete:
        reasons.append('report_window_incomplete')
    if not data_fresh:
        reasons.append('data_not_fresh')
    if int(sync_failed_count or 0) > 0:
        warnings.append('sync_has_failures')
    if metric.bind_event_count <= 0:
        warnings.append('no_success_bind_events')
    if not metric.is_dedupe_confirmed:
        warnings.append('dedupe_not_confirmed')
    if metric.duplicate_event_id_count:
        reasons.append('duplicate_event_id')
    if metric.duplicate_bind_id_count:
        warnings.append('duplicate_bind_id')
    if metric.duplicate_customer_user_id_count:
        warnings.append('duplicate_customer_user_id')
    if metric.missing_country_count:
        warnings.append('missing_country')
    if metric.missing_project_count:
        warnings.append('missing_project')
    if metric.missing_attribution_count:
        warnings.append('missing_attribution')
    if attribution_coverage is not None and float(attribution_coverage) < float(min_attribution_coverage):
        reasons.append('attribution_coverage_low')
    status = 'PASS'
    if reasons:
        status = 'BLOCKED'
    elif warnings:
        status = 'WARNING'
    return DataQualityGateResult(
        status=status,
        status_zh={'PASS': '通过', 'WARNING': '预警', 'BLOCKED': '阻断'}[status],
        reasons=sorted(set(reasons)),
        warnings=sorted(set(warnings)),
        checked_at_utc=datetime.now(timezone.utc).isoformat(),
    )


def fixture_ad_rows(report_date: date) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for scenario in FixtureRealConversionProvider.FIXTURE_SCENARIOS:
        spend = float(scenario['spend'])
        installs = max(float(scenario.get('binds') or 0) * 2.2, 1.0)
        clicks = max(installs * 4.0, 10.0)
        impressions = max(clicks / 0.035, 1000.0)
        row_date = report_date.isoformat()
        rows.append({
            'date': row_date,
            'platform': 'Meta',
            'account': 'fixture_meta_account',
            'account_id': 'fixture_meta_account',
            'country': scenario['country'],
            'campaign': scenario['campaign'],
            'ad_group': scenario['ad_group'],
            'ad': scenario['ad'],
            'cost': spend,
            'spend': spend,
            'impressions': impressions,
            'clicks': clicks,
            'link_clicks': clicks,
            'ctr': clicks / impressions,
            'cpm': spend / impressions * 1000 if impressions else 0.0,
            'installs': installs,
            'meta_installs': installs,
            'guild_joins': float(scenario.get('binds') or 0),
            'cpi': spend / installs if installs else None,
            'join_cost': spend / float(scenario['binds']) if float(scenario.get('binds') or 0) else None,
            'first_effective_spend_date': (report_date - timedelta(days=int(scenario.get('first_day_offset') or 1))).isoformat(),
            'frontend_trend': {
                'cpi_increase_pct': float(scenario.get('cpi_increase_pct') or 0),
                'ctr_decrease_pct': float(scenario.get('ctr_decrease_pct') or 0),
                'cpm_increase_pct': float(scenario.get('cpm_increase_pct') or 0),
            },
            'has_inner_winner': bool(scenario.get('has_inner_winner')),
            'mixed_change': bool(scenario.get('mixed_change')),
            'scenario': scenario['scenario'],
        })
    return rows


def _dashboard_detail_rows(snapshot: Dict[str, Any], report_date: date) -> List[Dict[str, Any]]:
    detail = snapshot.get('platform_detail_rows') or {}
    rows: List[Dict[str, Any]] = []
    for platform, platform_rows in detail.items():
        for row in platform_rows or []:
            candidate = dict(row or {})
            candidate.setdefault('platform', platform)
            candidate.setdefault('date', report_date.isoformat())
            candidate.setdefault('spend', candidate.get('cost'))
            candidate.setdefault('account_id', candidate.get('account'))
            rows.append(candidate)
    return rows


def _report_country_key(row: Dict[str, Any], fields: Tuple[str, ...]) -> Tuple[str, ...]:
    values: List[str] = []
    for field_name in fields:
        if field_name == 'account':
            raw = row.get('account_id') or row.get('account') or row.get('app_id')
        elif field_name == 'ad_group':
            raw = row.get('ad_group') or row.get('adset')
        elif field_name == 'ad':
            raw = row.get('ad') or row.get('ad_name')
        else:
            raw = row.get(field_name)
        values.append(str(raw or '').strip().lower())
    return tuple(values)


def _single_country_index(rows: List[Dict[str, Any]], fields: Tuple[str, ...]) -> Dict[Tuple[str, ...], str]:
    buckets: Dict[Tuple[str, ...], set[str]] = {}
    for row in rows or []:
        country = normalize_country_code((row or {}).get('country'))
        if country_is_unresolved(country):
            continue
        key = _report_country_key(row, fields)
        if not any(key):
            continue
        buckets.setdefault(key, set()).add(country)
    return {key: next(iter(countries)) for key, countries in buckets.items() if len(countries) == 1}


def enrich_report_row_countries(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    exact_index = _single_country_index(rows, ('platform', 'account', 'campaign', 'ad_group', 'ad'))
    group_index = _single_country_index(rows, ('platform', 'account', 'campaign', 'ad_group'))
    campaign_index = _single_country_index(rows, ('platform', 'account', 'campaign'))
    account_index = _single_country_index(rows, ('platform', 'account'))
    enriched: List[Dict[str, Any]] = []
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        country = normalize_country_code(row.get('country'))
        if not country_is_unresolved(country):
            enriched.append(row)
            continue
        inferred = (
            exact_index.get(_report_country_key(row, ('platform', 'account', 'campaign', 'ad_group', 'ad')))
            or group_index.get(_report_country_key(row, ('platform', 'account', 'campaign', 'ad_group')))
            or campaign_index.get(_report_country_key(row, ('platform', 'account', 'campaign')))
            or account_index.get(_report_country_key(row, ('platform', 'account')))
            or infer_country_from_ad_context(
                row.get('country'),
                row.get('account_id'),
                row.get('account'),
                row.get('app_id'),
                row.get('campaign'),
                row.get('ad_group'),
                row.get('adset'),
                row.get('ad'),
                row.get('ad_name'),
                row.get('project'),
                row.get('external_app'),
            )
        )
        updated = dict(row)
        updated['country'] = inferred or UNRESOLVED_COUNTRY_LABEL
        if inferred:
            warnings = list(updated.get('data_quality_warnings') or [])
            warnings.append('country_inferred_from_ad_context')
            updated['data_quality_warnings'] = warnings
        enriched.append(updated)
    return enriched


def _event_key(event: RealBindEvent) -> Tuple[str, str, str, str, str]:
    return (
        normalize_country_code(event.country),
        str(event.account_id or ''),
        str(event.campaign or ''),
        str(event.ad_group or ''),
        str(event.ad or ''),
    )


def _dashboard_real_bind_count(row: Dict[str, Any]) -> int:
    # Dashboard fact rows already reserve guild_joins for TUGAO true joins;
    # AF model callbacks live in af_guild_joins and must not drive advice.
    for key in ('real_bind_count', 'tugao_real_bind_count', 'guild_joins'):
        raw = row.get(key)
        if raw in (None, ''):
            continue
        try:
            return max(0, int(round(float(raw))))
        except (TypeError, ValueError):
            continue
    return 0


def _dashboard_real_bind_fact_available(row: Dict[str, Any]) -> bool:
    # A persisted zero is a valid TUGAO observation, not a missing source.
    return any(
        key in row and row.get(key) not in (None, '')
        for key in ('real_bind_count', 'tugao_real_bind_count', 'guild_joins')
    )


def _row_float(row: Dict[str, Any], *keys: str) -> float:
    for key in keys:
        raw = row.get(key)
        if raw in (None, ''):
            continue
        try:
            return max(0.0, float(raw))
        except (TypeError, ValueError):
            continue
    return 0.0


def _row_metric_availability(row: Dict[str, Any], *keys: str, source: str = 'observed') -> Dict[str, Any]:
    for key in keys:
        if key not in row:
            continue
        raw = row.get(key)
        if raw in (None, ''):
            continue
        try:
            value = max(0.0, float(raw))
        except (TypeError, ValueError):
            continue
        return {
            'value': value,
            'source': source,
            'is_available': True,
            'usable_for_rule': True,
            'field': key,
        }
    return {
        'value': 0.0,
        'source': 'missing',
        'is_available': False,
        'usable_for_rule': False,
        'field': '',
    }


def _derived_metric_availability(*dependencies: Dict[str, Any], value: float = 0.0) -> Dict[str, Any]:
    if dependencies and all(dep.get('is_available') for dep in dependencies):
        return {
            'value': max(0.0, float(value or 0.0)),
            'source': 'derived',
            'is_available': True,
            'usable_for_rule': True,
            'field': '',
        }
    return {
        'value': 0.0,
        'source': 'missing',
        'is_available': False,
        'usable_for_rule': False,
        'field': '',
    }


def _safe_rate(numerator: float, denominator: float) -> Optional[float]:
    if not denominator:
        return None
    return round(float(numerator or 0.0) / float(denominator), 6)


def _safe_cost(spend: float, count: float) -> Optional[float]:
    if not count:
        return None
    return round(float(spend or 0.0) / float(count), 4)


def _cap_im_child_metric(value: float, im_entries: float) -> float:
    if not im_entries:
        return 0.0
    return min(float(value or 0.0), float(im_entries or 0.0))


def _user_engaged_im_metric(
    *,
    first_reply: float,
    msg_ge_3: float,
    msg_ge_5: float,
    auto_apply_user_count: float,
) -> Tuple[float, str]:
    cap = float(auto_apply_user_count or 0.0)
    if not cap:
        return 0.0, 'missing_apply_user_count'
    if float(msg_ge_3 or 0.0) > 0:
        return _cap_im_child_metric(msg_ge_3, cap), 'im_ge_3'
    if float(first_reply or 0.0) > 0:
        return _cap_im_child_metric(first_reply, cap), 'first_user_reply_fallback'
    if float(msg_ge_5 or 0.0) > 0:
        return _cap_im_child_metric(msg_ge_5, cap), 'im_ge_5_fallback'
    return 0.0, 'im_ge_3'


def _row_funnel_metrics(row: Dict[str, Any], *, spend: float, real_binds: int) -> Dict[str, Any]:
    registrations_availability = _row_metric_availability(row, 'registrations', 'registration_count', 'onsite_registrations', 'onsite_registers', '新增注册', '站内注册')
    high_value_availability = _row_metric_availability(row, 'high_value_users', 'high_value', 'high_value_count', '高价值', '高价值用户')
    explicit_apply_availability = _row_metric_availability(row, 'auto_apply_user_count', 'apply_user_count', 'user_apply_count', '用户报名人数', '报名人数')
    legacy_apply_availability = _row_metric_availability(row, 'im_entries', 'enter_im_users', 'enter_im', '进入IM', '进入 IM', '进入IM人数')
    auto_apply_availability = _row_metric_availability(row, 'auto_apply_message_sent_count', 'auto_apply_message_users', 'auto_signup_users', 'auto_apply_users', '自动报名消息发送', '自动报名消息发送人数')
    first_reply_availability = _row_metric_availability(row, 'user_first_reply_users', 'im_first_reply_users', 'im_first_replies', 'IM首回', '用户首回人数', '首回人数')
    msg_ge_3_availability = _row_metric_availability(row, 'im_user_message_ge_3_users', 'im_manual_reply_3', '真人消息>=3', '真人消息>=3条', 'IM真人消息>=3', 'IM真人消息>=3条')
    msg_ge_5_availability = _row_metric_availability(row, 'im_user_message_ge_5_users', 'im_manual_reply_5', '真人消息>=5', '真人消息>=5条', 'IM真人消息>=5', 'IM真人消息>=5条')
    link_click_availability = _row_metric_availability(row, 'link_click_users', 'im_link_click_users', 'linky_link_click_users', '注册链接点击', 'IM链接点击', '链接点击人数')
    linky_register_availability = _row_metric_availability(row, 'linky_register_users', 'linky_registrations', 'linky_register_count', 'Linky注册', 'Linky 注册人数')
    bind_success_availability = _row_metric_availability(row, 'bind_success_users', 'bind_success_count', 'bind成功', 'bind 成功人数')
    crm_succeed_availability = _row_metric_availability(row, 'crm_succeed_users', 'crm_succeed_count', 'CRM succeed', 'CRM成功', 'CRM succeed 数')
    registrations = _row_float(row, 'registrations', 'registration_count', 'onsite_registrations', 'onsite_registers', '新增注册', '站内注册')
    high_value_users = _row_float(row, 'high_value_users', 'high_value', 'high_value_count', '高价值', '高价值用户')
    explicit_apply_user_count = _row_float(row, 'auto_apply_user_count', 'apply_user_count', 'user_apply_count', '用户报名人数', '报名人数')
    legacy_apply_user_count = _row_float(row, 'im_entries', 'enter_im_users', 'enter_im', '进入IM', '进入 IM', '进入IM人数')
    auto_apply_user_count = explicit_apply_user_count or legacy_apply_user_count
    auto_apply_user_count_source = 'auto_apply_user_count' if explicit_apply_user_count else ('legacy_im_entries' if legacy_apply_user_count else 'missing')
    auto_apply = _row_float(row, 'auto_apply_message_sent_count', 'auto_apply_message_users', 'auto_signup_users', 'auto_apply_users', '自动报名消息发送', '自动报名消息发送人数')
    first_reply = _row_float(row, 'user_first_reply_users', 'im_first_reply_users', 'IM首回', '用户首回人数', '首回人数')
    msg_ge_3 = _row_float(row, 'im_user_message_ge_3_users', 'im_manual_reply_3', '真人消息>=3', '真人消息>=3条', 'IM真人消息>=3', 'IM真人消息>=3条')
    msg_ge_5 = _row_float(row, 'im_user_message_ge_5_users', 'im_manual_reply_5', '真人消息>=5', '真人消息>=5条', 'IM真人消息>=5', 'IM真人消息>=5条')
    link_clicks = _row_float(row, 'link_click_users', 'im_link_click_users', 'linky_link_click_users', '注册链接点击', 'IM链接点击', '链接点击人数')
    linky_registers = _row_float(row, 'linky_register_users', 'linky_registrations', 'linky_register_count', 'Linky注册', 'Linky 注册人数')
    bind_success = _row_float(row, 'bind_success_users', 'bind_success_count', 'bind成功', 'bind 成功人数')
    crm_succeed = _row_float(row, 'crm_succeed_users', 'crm_succeed_count', 'CRM succeed', 'CRM成功', 'CRM succeed 数')

    system_touched = auto_apply
    user_engaged, user_engaged_metric_version = _user_engaged_im_metric(
        first_reply=first_reply,
        msg_ge_3=msg_ge_3,
        msg_ge_5=msg_ge_5,
        auto_apply_user_count=auto_apply_user_count,
    )
    user_engaged_availability = _derived_metric_availability(
        explicit_apply_availability if explicit_apply_availability.get('is_available') else legacy_apply_availability,
        msg_ge_3_availability,
        value=user_engaged,
    )
    high_intent = _cap_im_child_metric(max(link_clicks, linky_registers, bind_success), auto_apply_user_count)
    high_intent_availability = _derived_metric_availability(
        explicit_apply_availability if explicit_apply_availability.get('is_available') else legacy_apply_availability,
        link_click_availability,
        linky_register_availability,
        bind_success_availability,
        value=high_intent,
    )
    return {
        'registrations': registrations,
        'high_value_users': high_value_users,
        'im_entries': auto_apply_user_count,
        'auto_apply_user_count': auto_apply_user_count,
        'auto_apply_user_count_source': auto_apply_user_count_source,
        'auto_apply_message_users': auto_apply,
        'user_first_reply_users': first_reply,
        'im_user_message_ge_3_users': msg_ge_3,
        'im_user_message_ge_5_users': msg_ge_5,
        'link_click_users': link_clicks,
        'linky_register_users': linky_registers,
        'bind_success_users': bind_success,
        'crm_succeed_users': crm_succeed,
        'system_touched_im_users': system_touched,
        'user_engaged_im_users': user_engaged,
        'user_engaged_im_metric_version': user_engaged_metric_version,
        'high_intent_im_users': high_intent,
        'high_value_rate': _safe_rate(high_value_users, registrations or auto_apply_user_count),
        'registration_to_apply_rate': _safe_rate(auto_apply_user_count, registrations),
        'im_cost': _safe_cost(spend, auto_apply_user_count),
        'system_touched_im_rate': _safe_rate(system_touched, auto_apply_user_count),
        'user_engaged_im_rate': _safe_rate(user_engaged, auto_apply_user_count),
        'user_engaged_im_cost': _safe_cost(spend, user_engaged),
        'high_intent_im_rate': _safe_rate(high_intent, auto_apply_user_count),
        'high_intent_im_cost': _safe_cost(spend, high_intent),
        'im_to_join_rate': _safe_rate(real_binds, auto_apply_user_count),
        'linky_register_rate_from_link_click': _safe_rate(linky_registers, link_clicks),
        'bind_rate_from_linky': _safe_rate(bind_success, linky_registers),
        'crm_succeed_rate_from_bind': _safe_rate(crm_succeed, bind_success),
        'metric_availability': {
            'spend': {'value': float(spend or 0.0), 'source': 'observed', 'is_available': True, 'usable_for_rule': True, 'field': 'spend'},
            'registrations': registrations_availability,
            'high_value_users': high_value_availability,
            'auto_apply_user_count': explicit_apply_availability if explicit_apply_availability.get('is_available') else legacy_apply_availability,
            'auto_apply_message_users': auto_apply_availability,
            'user_first_reply_users': first_reply_availability,
            'im_user_message_ge_3_users': msg_ge_3_availability,
            'im_user_message_ge_5_users': msg_ge_5_availability,
            'user_engaged_im_users': user_engaged_availability,
            'link_click_users': link_click_availability,
            'linky_register_users': linky_register_availability,
            'bind_success_users': bind_success_availability,
            'crm_succeed_users': crm_succeed_availability,
            'real_join': {'value': float(real_binds or 0), 'source': 'observed', 'is_available': True, 'usable_for_rule': True, 'field': 'real_bind_count'},
            'high_intent_im_users': high_intent_availability,
        },
    }


def adapt_dashboard_snapshot_to_report_objects(
    snapshot: Dict[str, Any],
    events: List[RealBindEvent],
    report_date: date,
    *,
    allow_fixture_fallback: bool = True,
) -> List[AdObjectMetrics]:
    rows = enrich_report_row_countries(_dashboard_detail_rows(snapshot, report_date))
    if not rows and allow_fixture_fallback:
        rows = fixture_ad_rows(report_date)
    deduped = dedupe_first_effective_bind_events(events)
    bind_counts: Dict[Tuple[str, str, str, str, str], int] = {}
    for event in deduped:
        bind_counts[_event_key(event)] = bind_counts.get(_event_key(event), 0) + 1
    objects: List[AdObjectMetrics] = []
    for row in rows:
        country = normalize_country_code(row.get('country'))
        account_id = str(row.get('account_id') or row.get('account') or '').removeprefix('act_')
        account_identity = str(row.get('account_identity') or row.get('account') or account_id)
        campaign = str(row.get('campaign') or '')
        ad_group = str(row.get('ad_group') or row.get('adset') or '')
        ad = str(row.get('ad') or row.get('ad_name') or '')
        source_type = str(row.get('source_type') or '')
        key = (country, account_id, campaign, ad_group, ad)
        legacy_key = (country, account_identity, campaign, ad_group, ad)
        spend = float(row.get('spend') if row.get('spend') is not None else row.get('cost') or 0.0)
        installs = float(row.get('installs') or row.get('meta_installs') or 0.0)
        impressions = float(row.get('impressions') or 0.0)
        clicks = float(row.get('clicks') or row.get('link_clicks') or 0.0)
        is_fixture_row = bool(row.get('scenario'))
        real_binds = int(bind_counts.get(key, 0) or bind_counts.get(legacy_key, 0))
        row_real_binds = _dashboard_real_bind_count(row)
        row_has_real_bind_fact = _dashboard_real_bind_fact_available(row)
        attribution_quality = 'fixture' if is_fixture_row else ('tugao_raw_event' if real_binds else 'unknown')
        if not real_binds and is_fixture_row:
            real_binds = int(row.get('guild_joins') or 0)
        elif not is_fixture_row and not real_binds and row_has_real_bind_fact:
            real_binds = row_real_binds
            attribution_quality = 'tugao_funnel_fact'
        funnel_metrics = _row_funnel_metrics(row, spend=spend, real_binds=real_binds)
        first_spend = str(row.get('first_effective_spend_date') or snapshot.get('date_start') or report_date.isoformat())
        try:
            maturity_day = max(0, (report_date - datetime.strptime(first_spend[:10], '%Y-%m-%d').date()).days)
        except Exception:
            maturity_day = 0
        # Account labels were historically part of the stable object identity.
        # Preserve that identity while exposing numeric account_id for policy,
        # filtering and execution boundaries.
        object_id = _stable_id(country, account_identity, campaign, ad_group, ad)
        observation_identity = ad_observation_identity(
            platform=row.get('platform'),
            account_id=account_identity,
            country=country,
            campaign=campaign,
            ad_group=ad_group,
            ad=ad,
            source_type=source_type,
        )
        is_unresolved_country = country_is_unresolved(country)
        data_quality_warnings = list(row.get('data_quality_warnings') or [])
        if not account_id:
            data_quality_warnings.append('缺少广告账户信息')
        if is_unresolved_country:
            data_quality_warnings.append('无法识别投放国家，需要补齐 Meta/TUGAO 国家归属')
        objects.append(AdObjectMetrics(
            object_id=object_id,
            object_level='ad',
            country=country,
            project=str(row.get('project') or row.get('app_id') or 'unknown_project'),
            target_app=str(row.get('target_app') or 'inactive'),
            account_id=account_id,
            campaign=campaign,
            ad_group=ad_group,
            ad=ad,
            spend=round(spend, 4),
            impressions=round(impressions, 4),
            clicks=round(clicks, 4),
            ctr=round(float(row.get('ctr') or (clicks / impressions if impressions else 0.0)), 6),
            cpm=round(float(row.get('cpm') or (spend / impressions * 1000 if impressions else 0.0)), 4),
            installs=round(installs, 4),
            cpi=round(spend / installs, 4) if installs else None,
            registrations=round(float(funnel_metrics['registrations'] or 0.0), 4),
            high_value_users=round(float(funnel_metrics['high_value_users'] or 0.0), 4),
            im_entries=round(float(funnel_metrics['im_entries'] or 0.0), 4),
            auto_apply_user_count=round(float(funnel_metrics['auto_apply_user_count'] or 0.0), 4),
            auto_apply_message_users=round(float(funnel_metrics['auto_apply_message_users'] or 0.0), 4),
            auto_apply_user_count_source=str(funnel_metrics.get('auto_apply_user_count_source') or ''),
            user_first_reply_users=round(float(funnel_metrics['user_first_reply_users'] or 0.0), 4),
            im_user_message_ge_3_users=round(float(funnel_metrics['im_user_message_ge_3_users'] or 0.0), 4),
            im_user_message_ge_5_users=round(float(funnel_metrics['im_user_message_ge_5_users'] or 0.0), 4),
            link_click_users=round(float(funnel_metrics['link_click_users'] or 0.0), 4),
            linky_register_users=round(float(funnel_metrics['linky_register_users'] or 0.0), 4),
            bind_success_users=round(float(funnel_metrics['bind_success_users'] or 0.0), 4),
            crm_succeed_users=round(float(funnel_metrics['crm_succeed_users'] or 0.0), 4),
            system_touched_im_users=round(float(funnel_metrics['system_touched_im_users'] or 0.0), 4),
            user_engaged_im_users=round(float(funnel_metrics['user_engaged_im_users'] or 0.0), 4),
            high_intent_im_users=round(float(funnel_metrics['high_intent_im_users'] or 0.0), 4),
            high_value_rate=funnel_metrics['high_value_rate'],
            registration_to_apply_rate=funnel_metrics['registration_to_apply_rate'],
            im_cost=funnel_metrics['im_cost'],
            system_touched_im_rate=funnel_metrics['system_touched_im_rate'],
            user_engaged_im_rate=funnel_metrics['user_engaged_im_rate'],
            user_engaged_im_cost=funnel_metrics['user_engaged_im_cost'],
            user_engaged_im_metric_version=str(funnel_metrics.get('user_engaged_im_metric_version') or 'unknown'),
            high_intent_im_rate=funnel_metrics['high_intent_im_rate'],
            high_intent_im_cost=funnel_metrics['high_intent_im_cost'],
            im_to_join_rate=funnel_metrics['im_to_join_rate'],
            linky_register_rate_from_link_click=funnel_metrics['linky_register_rate_from_link_click'],
            bind_rate_from_linky=funnel_metrics['bind_rate_from_linky'],
            crm_succeed_rate_from_bind=funnel_metrics['crm_succeed_rate_from_bind'],
            af_guild_joins=float(row.get('af_guild_joins') or 0.0),
            real_bind_count=real_binds,
            real_bind_cpa=round(spend / real_binds, 4) if spend > 0 and real_binds else None,
            first_effective_spend_date=first_spend[:10],
            maturity_day=maturity_day,
            data_quality=DataQualityStatus(
                status='review' if not account_id or is_unresolved_country else 'ok',
                attribution_quality=attribution_quality,
                warnings=data_quality_warnings,
            ),
            budget_mode=BudgetMode(
                mode='mixed_change' if row.get('mixed_change') else 'unknown',
                cooldown_hours_remaining=72 if row.get('mixed_change') else 0,
            ),
            frontend_trend=dict(row.get('frontend_trend') or {}),
            metric_availability=dict(funnel_metrics.get('metric_availability') or {}),
            source_type=source_type,
            observation_identity=observation_identity,
        ))
    natural_indexes: Dict[str, int] = {}
    collapsed: List[AdObjectMetrics] = []
    for item in objects:
        source_type = str(item.source_type or '').strip().lower()
        if not any(marker in source_type for marker in ('自然', 'organic')):
            collapsed.append(item)
            continue
        existing_index = natural_indexes.get(item.object_id)
        if existing_index is None:
            natural_indexes[item.object_id] = len(collapsed)
            collapsed.append(item)
            continue
        existing = collapsed[existing_index]
        winner = item if int(item.real_bind_count or 0) > int(existing.real_bind_count or 0) else existing
        collapsed[existing_index] = replace(
            winner,
            source_type='自然量（多源去重）',
            real_bind_cpa=None,
        )
    return collapsed


class AdDailyRecommendationEngine:
    def __init__(self, config: Optional[Dict[str, Any]] = None) -> None:
        self.config = config or DEFAULT_RULE_CONFIG

    def market_cap(self, country: str) -> Optional[float]:
        market = (self.config.get('markets') or {}).get(normalize_country_code(country)) or {}
        cap = market.get('real_bind_cpa_cap')
        return None if cap is None else float(cap)

    def funnel_baseline(self, country: str) -> Dict[str, float]:
        baselines = self.config.get('rules', {}).get('funnel_baselines') or self.config.get('funnel_baselines') or {}
        return dict(baselines.get(normalize_country_code(country)) or baselines.get('default') or {})

    def sample_thresholds(self, item: AdObjectMetrics) -> Dict[str, float]:
        thresholds = (self.config.get('rules', {}).get('sample_thresholds') or self.config.get('sample_thresholds') or {})
        country = normalize_country_code(item.country)
        return {
            **dict(thresholds.get('default') or {}),
            **dict(thresholds.get(country) or {}),
        }

    def has_funnel_data(self, item: AdObjectMetrics) -> bool:
        return any(float(value or 0.0) > 0 for value in (
            item.im_entries,
            item.auto_apply_message_users,
            item.user_engaged_im_users,
            item.high_intent_im_users,
            item.link_click_users,
            item.linky_register_users,
            item.bind_success_users,
            item.crm_succeed_users,
        ))

    def metric_available(self, item: AdObjectMetrics, metric: str) -> bool:
        info = (item.metric_availability or {}).get(metric) or {}
        return bool(info.get('is_available') and info.get('usable_for_rule') and info.get('source') != 'missing')

    def metric_missing(self, item: AdObjectMetrics, metric: str) -> bool:
        return not self.metric_available(item, metric)

    def required_funnel_metrics_missing(self, item: AdObjectMetrics) -> List[str]:
        missing: List[str] = []
        target_app = str(getattr(item, 'target_app', '') or '').strip().lower()
        if target_app in {'linky', 'timo'}:
            for metric in ('auto_apply_user_count', 'user_engaged_im_users', 'link_click_users', 'linky_register_users', 'bind_success_users', 'crm_succeed_users'):
                if self.metric_missing(item, metric):
                    missing.append(metric)
        return missing

    def view_capability_missing(self, item: AdObjectMetrics) -> bool:
        target_app = str(getattr(item, 'target_app', '') or '').strip().lower()
        if target_app in {'linky', 'timo'} and self.required_funnel_metrics_missing(item):
            return True
        if target_app in {'linky', 'timo'} and not self.has_funnel_data(item) and (float(item.spend or 0.0) >= 5.0 or int(item.real_bind_count or 0) > 0):
            return True
        return False

    def post_funnel_event_inconsistent(self, item: AdObjectMetrics) -> bool:
        return (
            float(item.bind_success_users or 0.0) > 0
            and float(item.link_click_users or 0.0) <= 0
            and float(item.linky_register_users or 0.0) <= 0
            and self.metric_available(item, 'bind_success_users')
            and self.metric_available(item, 'link_click_users')
            and self.metric_available(item, 'linky_register_users')
        )

    def funnel_metrics_payload(self, item: AdObjectMetrics) -> Dict[str, Any]:
        return {
            'target_app': item.target_app,
            'registrations': item.registrations,
            'high_value_users': item.high_value_users,
            'high_value_rate': item.high_value_rate,
            'registration_to_apply_rate': item.registration_to_apply_rate,
            'im_entries': item.im_entries,
            'auto_apply_user_count': item.auto_apply_user_count,
            'auto_apply_user_count_source': item.auto_apply_user_count_source,
            'auto_apply_message_sent_count': item.auto_apply_message_users,
            'im_cost': item.im_cost,
            'system_touched_im_users': item.system_touched_im_users,
            'system_touched_im_rate': item.system_touched_im_rate,
            'user_engaged_im_users': item.user_engaged_im_users,
            'user_engaged_im_metric_version': item.user_engaged_im_metric_version,
            'user_engaged_im_rate': item.user_engaged_im_rate,
            'user_engaged_im_cost': item.user_engaged_im_cost,
            'high_intent_im_users': item.high_intent_im_users,
            'high_intent_im_rate': item.high_intent_im_rate,
            'high_intent_im_cost': item.high_intent_im_cost,
            'link_click_users': item.link_click_users,
            'linky_register_users': item.linky_register_users,
            'bind_success_users': item.bind_success_users,
            'crm_succeed_users': item.crm_succeed_users,
            'im_to_join_rate': item.im_to_join_rate,
            'linky_register_rate_from_link_click': item.linky_register_rate_from_link_click,
            'bind_rate_from_linky': item.bind_rate_from_linky,
            'crm_succeed_rate_from_bind': item.crm_succeed_rate_from_bind,
            'metric_availability': item.metric_availability,
        }

    def maturity_status(self, item: AdObjectMetrics, thresholds: Dict[str, float]) -> str:
        if int(item.maturity_day or 0) <= 0:
            return 'front_funnel_ready'
        min_im = float(thresholds.get('min_im') or 20)
        min_user_engaged = float(thresholds.get('min_user_engaged_im') or 10)
        min_join = float(thresholds.get('min_join') or 5)
        if float(item.im_entries or 0.0) < min_im:
            return 'front_funnel_ready'
        if float(item.user_engaged_im_users or 0.0) < min_user_engaged:
            return 'im_funnel_ready'
        if int(item.real_bind_count or 0) < min_join:
            return 'post_im_funnel_ready'
        return 'fully_matured'

    def action_gate(
        self,
        item: AdObjectMetrics,
        *,
        diagnosis_type: str,
        action_type: str,
        allow_generate_creative: bool,
        allow_pause: bool,
        allow_scale: bool,
        allow_reduce_budget: bool,
    ) -> Dict[str, Any]:
        scorecard = score_ad_object(item, country=normalize_country_code(item.country))
        stop_loss_guard = bool((scorecard.get('guardrails') or {}).get('stop_loss_candidate'))
        guards = dict(scorecard.get('guardrails') or {})
        core_gate = item.data_quality.attribution_quality not in {'fixture', 'simulated'}
        if core_gate and (allow_scale or allow_pause or allow_reduce_budget):
            scale_allowed = bool(allow_scale and guards.get('scale_candidate'))
            reduction_allowed = bool((allow_pause or allow_reduce_budget) and (guards.get('poor_candidate') or guards.get('stop_loss_candidate')))
            blocked_reasons = [] if scale_allowed or reduction_allowed else ['v4_core_guardrail_not_met']
            return {
                'allow_generate_creative': bool(allow_generate_creative and scale_allowed),
                'allow_pause': bool(allow_pause and reduction_allowed),
                'allow_scale': scale_allowed,
                'allow_reduce_budget': bool(allow_reduce_budget and reduction_allowed),
                'blocked_reasons': blocked_reasons,
            }
        min_strong_spend = float((self.config.get('rules') or {}).get('min_spend_for_strong_action') or 5.0)
        blocked_reasons: List[str] = []
        spend = float(item.spend or 0.0)
        target_app = str(getattr(item, 'target_app', '') or '').strip().lower()

        if spend < min_strong_spend:
            blocked_reasons.append('spend_below_minimum')
        missing = self.required_funnel_metrics_missing(item)
        if missing and not stop_loss_guard:
            blocked_reasons.append('required_metrics_missing')
        if target_app in {'linky', 'timo'} and self.view_capability_missing(item) and not stop_loss_guard:
            blocked_reasons.append('view_capability_missing')
        if self.post_funnel_event_inconsistent(item) and not stop_loss_guard:
            blocked_reasons.append('post_funnel_event_inconsistent')
        if item.data_quality.attribution_quality in {'fixture', 'simulated'}:
            blocked_reasons.append('fixture_or_test_data')
        apply_sample = float(item.auto_apply_user_count or item.im_entries or 0.0)
        if 0 < apply_sample < 20 and not stop_loss_guard:
            blocked_reasons.append('auto_apply_sample_insufficient')
        if item.im_cost == 0 or item.user_engaged_im_cost == 0 or item.real_bind_cpa == 0:
            blocked_reasons.append('invalid_zero_cost_evidence')
        if diagnosis_type in {'data_insufficient', 'data_anomaly', 'view_capability_missing', 'post_funnel_event_inconsistent', 'continue_observe', 'creative_effective_post_im_failed'}:
            blocked_reasons.append(diagnosis_type)
        if action_type in {'check_data_mapping', 'check_timo_im_mapping', 'check_linky_bind_crm_tracking', 'inspect_data_quality', 'inspect_post_im_funnel', 'inspect_business_result', 'observe'}:
            blocked_reasons.append('non_creative_action')

        deduped_reasons = list(dict.fromkeys(blocked_reasons))
        if not scorecard.get('strong_action_eligible') and not stop_loss_guard:
            blocked_reasons.append('v4_strong_action_guard_not_met')
        deduped_reasons = list(dict.fromkeys(blocked_reasons))
        blocked = bool(deduped_reasons)
        return {
            'allow_generate_creative': bool(allow_generate_creative and not blocked),
            'allow_pause': bool(allow_pause and not blocked),
            'allow_scale': bool(allow_scale and not blocked),
            'allow_reduce_budget': bool(allow_reduce_budget and not blocked),
            'blocked_reasons': deduped_reasons,
        }

    def _make_recommendation(
        self,
        item: AdObjectMetrics,
        window: Dict[str, str],
        *,
        cap: Optional[float],
        primary_action: str,
        reason_zh: str,
        status_tag: str,
        diagnosis_type: str,
        action_type: str,
        primary_layer: str,
        maturity_status: str,
        confidence: str = 'medium',
        adjustment_pct: int = 0,
        allow_pause: bool = False,
        allow_scale: bool = False,
        creative_scale_candidate: bool = False,
        business_scale_allowed: bool = False,
        allow_generate_creative: bool = False,
        evidence_points: Optional[List[str]] = None,
        needs_data: Optional[List[str]] = None,
        creative_diagnosis: Optional[Dict[str, Any]] = None,
        post_im_diagnosis: Optional[Dict[str, Any]] = None,
        business_diagnosis: Optional[Dict[str, Any]] = None,
    ) -> Recommendation:
        scorecard = score_ad_object(item, country=normalize_country_code(item.country))
        evidence = RecommendationEvidence(
            spend=item.spend,
            installs=item.installs,
            real_bind_count=int(item.real_bind_count or 0),
            real_bind_cpa=item.real_bind_cpa,
            cpi=item.cpi,
            ctr=item.ctr,
            cpm=item.cpm,
            country_cap=cap,
            data_window=window,
            data_quality=item.data_quality,
            budget_mode=item.budget_mode,
            frontend_trend=item.frontend_trend,
            funnel_metrics=self.funnel_metrics_payload(item),
            evidence_points=list(evidence_points or []),
            scorecard=scorecard,
        )
        name = item.ad or item.ad_group or item.campaign or item.object_id
        gate = self.action_gate(
            item,
            diagnosis_type=diagnosis_type,
            action_type=action_type,
            allow_generate_creative=allow_generate_creative,
            allow_pause=allow_pause,
            allow_scale=allow_scale,
            allow_reduce_budget=False,
        )
        return Recommendation(
            recommendation_id=_stable_id(item.observation_identity or item.object_id, primary_action, diagnosis_type, adjustment_pct, RECOMMENDATION_RULE_VERSION),
            object_id=item.object_id,
            object_level=item.object_level,
            object_name=name,
            country=item.country,
            project=item.project,
            primary_action=primary_action,
            primary_action_zh=ZH_LABELS.get(primary_action, primary_action),
            adjustment_pct=adjustment_pct,
            reason_zh=reason_zh,
            confidence=confidence,
            confidence_zh={'high': '高', 'medium': '中', 'low': '低'}.get(confidence, confidence),
            status_tag=status_tag,
            evidence=evidence,
            created_at_utc=datetime.now(timezone.utc).isoformat(),
            diagnosis_type=diagnosis_type,
            diagnosis_type_zh=ZH_LABELS.get(diagnosis_type, diagnosis_type),
            action_type=action_type,
            action_type_zh=ZH_LABELS.get(action_type, action_type),
            primary_layer=primary_layer,
            maturity_status=maturity_status,
            data_quality_status=item.data_quality.status,
            allow_pause=gate['allow_pause'],
            allow_scale=gate['allow_scale'],
            creative_scale_candidate=creative_scale_candidate,
            business_scale_allowed=bool(business_scale_allowed and gate['allow_scale']),
            allow_generate_creative=gate['allow_generate_creative'],
            needs_data=list(needs_data or []),
            creative_diagnosis=dict(creative_diagnosis or {}),
            post_im_diagnosis=dict(post_im_diagnosis or {}),
            business_diagnosis=dict(business_diagnosis or {}),
            action_gate=gate,
            decision_context={
                'evidence': {
                    'points': list(evidence_points or []),
                    'data_window': dict(window),
                    'rule_version': RECOMMENDATION_RULE_VERSION,
                },
                'risk': {
                    'blocked': bool(gate['blocked_reasons']),
                    'reasons': list(gate['blocked_reasons']),
                },
                'confidence': confidence,
                'recommended_action': primary_action,
                'diagnosis_type': diagnosis_type,
            },
            observation_identity=item.observation_identity,
        )

    def _diagnosis_payloads(
        self,
        item: AdObjectMetrics,
        *,
        baseline: Dict[str, Any],
        cap: Optional[float],
        cpa: Optional[float],
        im_cost_ok: bool,
        registration_to_apply_ok: bool,
        user_engaged_cost_ok: bool,
        user_engaged_rate_ok: bool,
        high_value_ok: bool,
        im_to_join_ok: bool,
        post_link_drop: bool,
        final_ok: bool,
        min_user_engaged: float,
    ) -> Tuple[Dict[str, Any], Dict[str, Any], Dict[str, Any]]:
        creative_evidence: List[str] = []
        if im_cost_ok:
            creative_evidence.append('用户报名成本正常')
        else:
            creative_evidence.append('用户报名成本偏高或缺失')
        if registration_to_apply_ok:
            creative_evidence.append('注册→用户报名率正常')
        else:
            creative_evidence.append('注册→用户报名率偏低或缺失')
        if user_engaged_cost_ok:
            creative_evidence.append('用户行为型有效 IM 成本正常')
        else:
            creative_evidence.append('用户行为型有效 IM 成本偏高或缺失')
        if user_engaged_rate_ok:
            creative_evidence.append('用户行为型有效 IM 率正常')
        else:
            creative_evidence.append('用户行为型有效 IM 率偏低')
        if high_value_ok:
            creative_evidence.append('高价值占比正常')
        else:
            creative_evidence.append('高价值占比偏低')
        if float(item.user_engaged_im_users or 0.0) < min_user_engaged:
            creative_status = 'weak'
        elif im_cost_ok and user_engaged_cost_ok and user_engaged_rate_ok and high_value_ok:
            creative_status = 'effective'
        else:
            creative_status = 'weak'

        post_evidence = [
            f'链接点击 {item.link_click_users:g}',
            f'Linky 注册 {item.linky_register_users:g}',
            f'bind 成功 {item.bind_success_users:g}',
            f'CRM succeed {item.crm_succeed_users:g}',
        ]
        if post_link_drop:
            post_status = 'weak'
        elif item.im_to_join_rate is not None and not im_to_join_ok:
            post_status = 'weak'
            post_evidence.append('IM 到入会转化低于警戒线')
        else:
            post_status = 'healthy'

        business_evidence = [
            f'真实入会 {int(item.real_bind_count or 0)}',
            f'真实入会成本 {"-" if cpa is None else f"${cpa:.2f}"}',
            f'国家红线 {"-" if cap is None else f"${cap:.2f}"}',
        ]
        if cap is None or cpa is None:
            business_status = 'unknown'
        elif final_ok:
            business_status = 'healthy'
        else:
            business_status = 'cpa_high'

        return (
            {
                'status': creative_status,
                'scope': 'front_funnel_and_user_engaged_im',
                'evidence': creative_evidence,
                'metrics': {
                    'auto_apply_user_count': item.auto_apply_user_count,
                    'auto_apply_user_count_source': item.auto_apply_user_count_source,
                    'auto_apply_message_sent_count': item.auto_apply_message_users,
                    'registration_to_apply_rate': item.registration_to_apply_rate,
                    'user_apply_cost': item.im_cost,
                    'user_engaged_im_users': item.user_engaged_im_users,
                    'user_engaged_im_metric_version': item.user_engaged_im_metric_version,
                    'user_engaged_im_rate': item.user_engaged_im_rate,
                    'user_engaged_im_cost': item.user_engaged_im_cost,
                    'high_value_rate': item.high_value_rate,
                },
            },
            {
                'status': post_status,
                'scope': 'post_im_funnel',
                'evidence': post_evidence,
                'metrics': {
                    'link_click_users': item.link_click_users,
                    'linky_register_users': item.linky_register_users,
                    'bind_success_users': item.bind_success_users,
                    'crm_succeed_users': item.crm_succeed_users,
                    'im_to_join_rate': item.im_to_join_rate,
                },
            },
            {
                'status': business_status,
                'scope': 'business_result',
                'evidence': business_evidence,
                'metrics': {
                    'real_bind_count': item.real_bind_count,
                    'real_bind_cpa': item.real_bind_cpa,
                    'country_cap': cap,
                    'spend': item.spend,
                },
            },
        )

    def recommend_with_funnel(
        self,
        item: AdObjectMetrics,
        window: Dict[str, str],
        *,
        data_quality_gate: Optional[DataQualityGateResult] = None,
    ) -> Recommendation:
        cap = self.market_cap(item.country)
        baseline = self.funnel_baseline(item.country)
        thresholds = self.sample_thresholds(item)
        maturity = self.maturity_status(item, thresholds)
        spend = float(item.spend or 0.0)
        binds = int(item.real_bind_count or 0)
        cpa = item.real_bind_cpa
        evidence_points = [
            f'用户报名 {item.auto_apply_user_count:g}，用户报名成本 {"-" if item.im_cost is None else f"${item.im_cost:.2f}"}',
            f'注册→用户报名率 {"-" if item.registration_to_apply_rate is None else f"{item.registration_to_apply_rate:.1%}"}',
            f'用户行为型有效 IM {item.user_engaged_im_users:g}，口径 {item.user_engaged_im_metric_version}，成本 {"-" if item.user_engaged_im_cost is None else f"${item.user_engaged_im_cost:.2f}"}',
            f'链接点击 {item.link_click_users:g}，Linky 注册 {item.linky_register_users:g}，bind 成功 {item.bind_success_users:g}',
            f'真实入会 {binds}，用户报名→入会率 {"-" if item.im_to_join_rate is None else f"{item.im_to_join_rate:.1%}"}',
        ]
        needs_data: List[str] = []
        if item.auto_apply_user_count_source in {'missing', ''}:
            needs_data.append('用户报名人数')
        elif item.auto_apply_user_count_source == 'legacy_im_entries':
            needs_data.append('独立 auto_apply_user_count 字段')
        if item.link_click_users and not item.linky_register_users:
            needs_data.append('Linky 注册人数')
        if item.linky_register_users and not item.bind_success_users:
            needs_data.append('bind 成功人数')
        if item.bind_success_users and not item.crm_succeed_users:
            needs_data.append('CRM succeed 数')
        if not item.user_first_reply_users:
            needs_data.append('用户首回人数')
        if item.user_engaged_im_metric_version != 'im_ge_3':
            needs_data.append('真人消息>=3人数')
        needs_data.append('1 分钟内响应率')

        scorecard = score_ad_object(item, country=normalize_country_code(item.country))
        use_v4 = item.data_quality.attribution_quality not in {'fixture', 'simulated'}

        if not use_v4 and data_quality_gate is not None and data_quality_gate.status == 'BLOCKED':
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh='数据质量闸门未通过，暂不输出暂停、放量或素材生成动作。',
                status_tag='data_missing', diagnosis_type='data_anomaly', action_type='inspect_data_quality',
                primary_layer='data_quality', maturity_status=maturity, confidence='low',
                evidence_points=list(data_quality_gate.reasons or []) + evidence_points,
                needs_data=needs_data,
            )
        if not use_v4 and item.data_quality.status != 'ok':
            return self._make_recommendation(
                item, window, cap=cap, primary_action='manual_review',
                reason_zh='广告对象数据质量不足，先人工复核归因、国家和账户映射。',
                status_tag='data_quality', diagnosis_type='data_anomaly', action_type='inspect_data_quality',
                primary_layer='data_quality', maturity_status=maturity, confidence='low',
                evidence_points=list(item.data_quality.warnings or []) + evidence_points,
                needs_data=needs_data,
            )
        min_strong_spend = float((self.config.get('rules') or {}).get('min_spend_for_strong_action') or 5.0)
        post_inconsistent = self.post_funnel_event_inconsistent(item)
        if not use_v4 and self.view_capability_missing(item):
            target_label = {'linky': 'Linky', 'timo': 'Timo'}.get(str(item.target_app or '').lower(), item.target_app or '当前')
            action_type = 'check_timo_im_mapping' if str(item.target_app or '').lower() == 'timo' else 'check_data_mapping'
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh=f'{target_label} 视图下 IM / 后链路指标未接入或未映射，本轮不能执行素材链路诊断。',
                status_tag='data_missing', diagnosis_type='view_capability_missing', action_type=action_type,
                primary_layer='view_capability', maturity_status=maturity, confidence='low',
                evidence_points=evidence_points,
                needs_data=needs_data + ['视图 IM/后链路指标映射'],
            )
        if not use_v4 and post_inconsistent and spend >= min_strong_spend:
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh='观测到 bind 成功但链接点击和 Linky 注册为 0，优先判定为后链路事件异常或归因断层，不允许反向判素材差。',
                status_tag='data_quality', diagnosis_type='post_funnel_event_inconsistent', action_type='check_linky_bind_crm_tracking',
                primary_layer='data_quality', maturity_status=maturity, confidence='low',
                evidence_points=evidence_points + ['post_funnel_inconsistent'],
                needs_data=needs_data + ['Linky 点击/注册/bind 事件对账'],
            )
        if not use_v4 and post_inconsistent and binds > 0 and float(item.auto_apply_user_count or item.im_entries or 0.0) <= 0:
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh='真实入会或 bind 成功存在，但用户报名、链接点击和 Linky 注册均为 0，优先判定为数据异常或事件回传断层。',
                status_tag='data_quality', diagnosis_type='data_anomaly', action_type='check_linky_bind_crm_tracking',
                primary_layer='data_quality', maturity_status=maturity, confidence='low',
                evidence_points=evidence_points + ['post_funnel_inconsistent', 'zero_apply_with_downstream_success'],
                needs_data=needs_data + ['用户报名/Linky 点击/注册/bind 事件对账'],
            )
        if use_v4:
            maturity = str(scorecard.get('band') or 'data_insufficient')
            core_labels = {
                'real_join_cpa': '入会单价', 'real_joins': '真实入会', 'cpi': '安装单价',
                'installs': '安装数', 'ctr': 'CTR',
            }
            dimension_risks = [
                core_labels.get(name, name) for name, detail in (scorecard.get('dimensions') or {}).items()
                if detail.get('available') and float(detail.get('score') or 0) < 45
            ]
            score_text = '-' if scorecard.get('score') is None else f"{float(scorecard['score']):.1f}"
            maturity_points = [
                f"{core_labels.get(name, name)} {detail.get('value', '-')}"
                + (f"/{detail.get('strong_threshold')}" if detail.get('strong_threshold') is not None else '')
                + f"（{detail.get('state')}）"
                for name, detail in (scorecard.get('maturity') or {}).items()
            ]
            score_points = [
                f"v4.2 得分 {score_text}，等级 {scorecard.get('band_zh')}",
            ] + maturity_points
            guards = scorecard.get('guardrails') or {}
            if guards.get('stop_loss_candidate'):
                return self._make_recommendation(
                    item, window, cap=cap, primary_action='pause', adjustment_pct=-100,
                    reason_zh='安装不少于 100 但真实入会仍为 0，且 CPI/CTR 出现风险，进入止损候选；仍需人工审批，默认不写 Meta。',
                    status_tag='zero_stop', diagnosis_type='business_result_anomaly', action_type='pause',
                    primary_layer='business_result', maturity_status=maturity, confidence=scorecard.get('confidence', 'medium'),
                    allow_pause=True, evidence_points=score_points, needs_data=needs_data,
                )
            if guards.get('poor_candidate'):
                return self._make_recommendation(
                    item, window, cap=cap, primary_action='reduce_budget', adjustment_pct=-20,
                    reason_zh='安装、真实入会和成本样本成熟，CPA 高于国家较差线，判为降预算候选；预算动作仍需人工审批。',
                    status_tag='over_cap', diagnosis_type='business_result_anomaly', action_type='reduce_budget',
                    primary_layer='business_result', maturity_status=maturity, confidence=scorecard.get('confidence', 'medium'),
                    allow_pause=True, evidence_points=score_points, needs_data=needs_data,
                )
            if guards.get('scale_candidate'):
                return self._make_recommendation(
                    item, window, cap=cap, primary_action='scale_up', adjustment_pct=10,
                    reason_zh='安装、真实入会和成本样本成熟，CPA 达优秀线且 CPI/CTR 达扩量门槛，进入扩量候选；仍需人工审批。',
                    status_tag='winner', diagnosis_type='scale_opportunity', action_type='scale_up',
                    primary_layer='business_result', maturity_status=maturity, confidence=scorecard.get('confidence', 'high'),
                    allow_scale=True, creative_scale_candidate=True, business_scale_allowed=True,
                    allow_generate_creative=True,
                    evidence_points=score_points, needs_data=needs_data,
                )
            if scorecard.get('band') == 'data_insufficient':
                business_result_available = scorecard.get('business_result_available') is True
                reason_zh = (
                    '真实入会数据已接入，当前样本尚未达到强动作门槛；系统继续积累样本，不据此调整预算。'
                    if business_result_available and binds == 0
                    else '真实入会结果尚未接入，或五项核心指标中存在关键缺项；系统继续补数，不要求用户核对。'
                )
                return self._make_recommendation(
                    item, window, cap=cap, primary_action='observe',
                    reason_zh=reason_zh,
                    status_tag='data_insufficient', diagnosis_type='data_insufficient', action_type='observe',
                    primary_layer='sample_maturity', maturity_status=maturity, confidence='low',
                    evidence_points=score_points, needs_data=needs_data,
                )
            risk_text = '、'.join(dimension_risks) if dimension_risks else '暂无单项强风险'
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh=f'业务结果已成熟，但未命中较差、止损或扩量强护栏；当前风险维度：{risk_text}，继续观察并分层修正。',
                status_tag='observe', diagnosis_type='continue_observe', action_type='observe',
                primary_layer='business_result', maturity_status=maturity, confidence=scorecard.get('confidence', 'medium'),
                evidence_points=score_points, needs_data=needs_data,
            )
        if spend < min_strong_spend:
            extra_points = ['post_funnel_inconsistent'] if post_inconsistent else []
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh=f'消耗 ${spend:.2f} 低于强动作门槛 ${min_strong_spend:.2f}，成本指标无效，只能继续观察。',
                status_tag='sample_insufficient', diagnosis_type='sample_insufficient', action_type='observe',
                primary_layer='sample_maturity', maturity_status=maturity, confidence='low',
                evidence_points=evidence_points + extra_points,
                needs_data=needs_data,
            )
        min_im = float(thresholds.get('min_im') or 20)
        min_apply_for_rate_diagnosis = float(thresholds.get('min_apply_for_rate_diagnosis') or 50)
        min_user_engaged = float(thresholds.get('min_user_engaged_im') or 10)
        min_join = float(thresholds.get('min_join') or 5)
        min_observation_hours = float(thresholds.get('min_observation_hours') or 24)
        observation_hours = int(item.maturity_day or 0) * 24
        has_enough_front_signal = (
            float(item.im_entries or 0.0) >= min_im
            or float(item.user_engaged_im_users or 0.0) >= min_user_engaged
            or binds > 0
        )
        if observation_hours < min_observation_hours and not has_enough_front_signal:
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh='观察窗口或进入 IM 样本不足，只做影子诊断，不触发强动作。',
                status_tag='sample_insufficient', diagnosis_type='sample_insufficient', action_type='observe',
                primary_layer='sample_maturity', maturity_status=maturity, confidence='low',
                evidence_points=evidence_points,
                needs_data=needs_data,
            )
        if float(item.im_entries or 0.0) < min_im and binds <= 0:
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh='进入 IM 样本不足且暂无真实入会，只做影子诊断，不触发强动作。',
                status_tag='sample_insufficient', diagnosis_type='sample_insufficient', action_type='observe',
                primary_layer='sample_maturity', maturity_status=maturity, confidence='low',
                evidence_points=evidence_points,
                needs_data=needs_data,
            )
        if 0 < float(item.auto_apply_user_count or item.im_entries or 0.0) < min_im:
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh='用户报名样本不足 20，不能判断素材质量，也不触发修素材。',
                status_tag='sample_insufficient', diagnosis_type='sample_insufficient', action_type='observe',
                primary_layer='sample_maturity', maturity_status=maturity, confidence='low',
                evidence_points=evidence_points,
                needs_data=needs_data,
            )

        im_cost_ok = item.im_cost is not None and item.im_cost <= float(baseline.get('im_cost_p50') or item.im_cost) * 1.15
        registration_to_apply_ok = item.registration_to_apply_rate is None or item.registration_to_apply_rate >= float(baseline.get('registration_to_apply_rate_normal') or 0.60)
        user_engaged_cost_available = float(item.user_engaged_im_users or 0.0) >= min_user_engaged
        user_engaged_cost_ok = (not user_engaged_cost_available) or (
            item.user_engaged_im_cost is not None
            and item.user_engaged_im_cost <= float(baseline.get('user_engaged_im_cost_p50') or item.user_engaged_im_cost) * 1.20
        )
        user_engaged_rate_ok = item.user_engaged_im_rate is not None and item.user_engaged_im_rate >= float(baseline.get('user_engaged_im_rate_p50') or 0.2) * 0.75
        user_engaged_rate_severely_low = (
            item.user_engaged_im_rate is not None
            and item.user_engaged_im_rate < float(baseline.get('user_engaged_im_rate_p50') or 0.2) * 0.50
        )
        high_value_ok = item.high_value_rate is None or item.high_value_rate >= float(baseline.get('high_value_rate_p50') or 0.5) * 0.75
        im_to_join_ok = item.im_to_join_rate is not None and item.im_to_join_rate >= float(baseline.get('im_to_join_rate_warning') or baseline.get('im_to_join_rate_p50') or 0.1)
        final_ok = cap is not None and cpa is not None and cpa <= cap
        post_link_drop = (
            (item.link_click_users and item.linky_register_users and item.linky_register_rate_from_link_click is not None and item.linky_register_rate_from_link_click < 0.35)
            or (item.linky_register_users and item.bind_success_users and item.bind_rate_from_linky is not None and item.bind_rate_from_linky < 0.50)
            or (item.bind_success_users and item.crm_succeed_users and item.crm_succeed_rate_from_bind is not None and item.crm_succeed_rate_from_bind < 0.80)
        )
        creative_diagnosis, post_im_diagnosis, business_diagnosis = self._diagnosis_payloads(
            item,
            baseline=baseline,
            cap=cap,
            cpa=cpa,
            im_cost_ok=im_cost_ok,
            registration_to_apply_ok=registration_to_apply_ok,
            user_engaged_cost_ok=user_engaged_cost_ok,
            user_engaged_rate_ok=user_engaged_rate_ok,
            high_value_ok=high_value_ok,
            im_to_join_ok=im_to_join_ok,
            post_link_drop=bool(post_link_drop),
            final_ok=final_ok,
            min_user_engaged=min_user_engaged,
        )
        creative_front_ok = im_cost_ok and registration_to_apply_ok and high_value_ok
        im_user_quality_ok = user_engaged_cost_ok and user_engaged_rate_ok and float(item.user_engaged_im_users or 0.0) >= min_user_engaged
        post_im_ok = post_im_diagnosis.get('status') == 'healthy'
        apply_sample = float(item.auto_apply_user_count or item.im_entries or 0.0)

        if min_im <= apply_sample < min_apply_for_rate_diagnosis and binds >= min_join:
            engaged_count = float(item.user_engaged_im_users or 0.0)
            engaged_maturity = (
                f'用户行为型有效 IM {engaged_count:g}/{min_user_engaged:g} 尚未成熟'
                if engaged_count < min_user_engaged
                else f'用户行为型有效 IM {engaged_count:g}/{min_user_engaged:g} 已成熟'
            )
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh=(
                    f'安装 {item.installs:g}、真实入会 {binds} 已达到结果样本门槛 {min_join:g}；'
                    f'用户报名 {apply_sample:g}/{min_apply_for_rate_diagnosis:g} 尚未达到素材强判门槛，'
                    f'{engaged_maturity}，因此暂不下素材因果结论，由系统继续观察。'
                ),
                status_tag='observe', diagnosis_type='continue_observe',
                action_type='observe', primary_layer='business_result',
                maturity_status=maturity, confidence='medium', allow_generate_creative=False,
                evidence_points=evidence_points + ['最终入会样本已成熟；用户报名与有效 IM 成熟度分开判断'],
                needs_data=needs_data,
                creative_diagnosis=creative_diagnosis,
                post_im_diagnosis=post_im_diagnosis,
                business_diagnosis=business_diagnosis,
            )

        if min_im <= apply_sample < min_apply_for_rate_diagnosis and not user_engaged_rate_ok:
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh='用户行为型有效 IM 偏弱，但用户报名样本尚未达到强判门槛，只记录风险并继续观察。',
                status_tag='sample_insufficient', diagnosis_type='sample_insufficient',
                action_type='observe', primary_layer='sample_maturity',
                maturity_status=maturity, confidence='low', allow_generate_creative=False,
                evidence_points=evidence_points,
                needs_data=needs_data,
                creative_diagnosis=creative_diagnosis,
                post_im_diagnosis=post_im_diagnosis,
                business_diagnosis=business_diagnosis,
            )

        if im_cost_ok and apply_sample >= min_apply_for_rate_diagnosis and user_engaged_rate_severely_low:
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh='用户报名分母已经足够，但用户行为型有效 IM 率严重偏低，倾向低质量流量或素材承诺偏差。',
                status_tag='low_quality_traffic', diagnosis_type='low_quality_traffic',
                action_type='generate_repair_creative', primary_layer='effective_im',
                maturity_status=maturity, confidence='medium', allow_generate_creative=True,
                evidence_points=evidence_points,
                needs_data=needs_data,
                creative_diagnosis=creative_diagnosis,
                post_im_diagnosis=post_im_diagnosis,
                business_diagnosis=business_diagnosis,
            )

        if not im_cost_ok or not user_engaged_cost_ok:
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh='素材前链路或用户行为型有效 IM 成本偏弱，允许生成修正素材，但不自动暂停。',
                status_tag='front_funnel_weak', diagnosis_type='front_funnel_weak',
                action_type='generate_repair_creative', primary_layer='front_funnel',
                maturity_status=maturity, confidence='medium', allow_generate_creative=True,
                evidence_points=evidence_points,
                needs_data=needs_data,
                creative_diagnosis=creative_diagnosis,
                post_im_diagnosis=post_im_diagnosis,
                business_diagnosis=business_diagnosis,
            )

        if im_cost_ok and not user_engaged_rate_ok:
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh='用户报名便宜但用户行为型有效率偏低，倾向点击诱导或素材承诺偏差。',
                status_tag='low_quality_traffic', diagnosis_type='low_quality_traffic',
                action_type='generate_repair_creative', primary_layer='effective_im',
                maturity_status=maturity, confidence='medium', allow_generate_creative=True,
                evidence_points=evidence_points,
                needs_data=needs_data,
                creative_diagnosis=creative_diagnosis,
                post_im_diagnosis=post_im_diagnosis,
                business_diagnosis=business_diagnosis,
            )

        if creative_front_ok and im_user_quality_ok and not post_im_ok:
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh='素材已带来可互动 IM 用户，但入会或 Linky/bind/CRM 后链路掉线；不建议直接重画素材。',
                status_tag='post_im_failed', diagnosis_type='creative_effective_post_im_failed',
                action_type='inspect_post_im_funnel', primary_layer='post_im_funnel',
                maturity_status=maturity, confidence='medium', allow_generate_creative=False,
                evidence_points=evidence_points,
                needs_data=needs_data,
                creative_diagnosis=creative_diagnosis,
                post_im_diagnosis=post_im_diagnosis,
                business_diagnosis=business_diagnosis,
            )

        if creative_front_ok and im_user_quality_ok and post_im_ok and not final_ok and cap is not None and cpa is not None:
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh='素材前链路、用户行为型 IM 和后链路均未见明显异常，但真实入会成本高于国家红线；先查归因、国家红线和成本结构，不触发修素材。',
                status_tag='business_result_anomaly', diagnosis_type='business_result_anomaly',
                action_type='inspect_business_result', primary_layer='business_result',
                maturity_status=maturity, confidence='medium', allow_generate_creative=False,
                allow_scale=False, allow_pause=False,
                evidence_points=evidence_points,
                needs_data=needs_data,
                creative_diagnosis=creative_diagnosis,
                post_im_diagnosis=post_im_diagnosis,
                business_diagnosis=business_diagnosis,
            )

        if creative_front_ok and im_user_quality_ok and post_im_ok and final_ok:
            pct = 10
            if binds >= int((self.config.get('rules') or {}).get('high_confidence_binds') or 20) and final_ok:
                pct = 15
            return self._make_recommendation(
                item, window, cap=cap, primary_action='scale_up',
                reason_zh='素材前链路、用户行为型 IM、后链路和最终经营校验均正常，建议放量并生成同方向衍生素材。',
                status_tag='winner', diagnosis_type='scale_opportunity',
                action_type='generate_derivative_creative', primary_layer='growth_opportunity',
                maturity_status=maturity, confidence='high' if final_ok and binds >= 10 else 'medium',
                adjustment_pct=pct, allow_scale=True, creative_scale_candidate=True,
                business_scale_allowed=True, allow_generate_creative=True,
                evidence_points=evidence_points,
                needs_data=needs_data,
                creative_diagnosis=creative_diagnosis,
                post_im_diagnosis=post_im_diagnosis,
                business_diagnosis=business_diagnosis,
            )

        if creative_front_ok and im_user_quality_ok and post_im_ok and im_to_join_ok:
            return self._make_recommendation(
                item, window, cap=cap, primary_action='observe',
                reason_zh='素材方向值得扩展，但真实入会 CPA 或国家经营红线口径尚未确认，暂不建议直接加预算。',
                status_tag='potential_winner', diagnosis_type='creative_scale_candidate',
                action_type='generate_derivative_creative', primary_layer='growth_opportunity',
                maturity_status=maturity, confidence='medium', allow_scale=False,
                creative_scale_candidate=True, business_scale_allowed=False,
                allow_generate_creative=True,
                evidence_points=evidence_points,
                needs_data=needs_data,
                creative_diagnosis=creative_diagnosis,
                post_im_diagnosis=post_im_diagnosis,
                business_diagnosis=business_diagnosis,
            )

        return self._make_recommendation(
            item, window, cap=cap, primary_action='observe',
            reason_zh='分段指标未触发强诊断，继续观察并等待更多后链路数据。',
            status_tag='observe', diagnosis_type='continue_observe', action_type='observe',
            primary_layer='observation', maturity_status=maturity, confidence='low',
            evidence_points=evidence_points,
            needs_data=needs_data,
            creative_diagnosis=creative_diagnosis,
            post_im_diagnosis=post_im_diagnosis,
            business_diagnosis=business_diagnosis,
        )

    def recommend(
        self,
        item: AdObjectMetrics,
        window: Dict[str, str],
        *,
        data_quality_gate: Optional[DataQualityGateResult] = None,
    ) -> Recommendation:
        min_strong_spend = float((self.config.get('rules') or {}).get('min_spend_for_strong_action') or 5.0)
        if (
            self.has_funnel_data(item)
            or self.view_capability_missing(item)
            or self.post_funnel_event_inconsistent(item)
            or float(item.spend or 0.0) < min_strong_spend
        ):
            return self.recommend_with_funnel(item, window, data_quality_gate=data_quality_gate)

        cap = self.market_cap(item.country)
        rules = self.config.get('rules') or {}
        binds = int(item.real_bind_count or 0)
        cpa = item.real_bind_cpa
        spend = float(item.spend or 0.0)
        maturity = int(item.maturity_day or 0)
        frontend = item.frontend_trend or {}
        risk_cfg = rules.get('frontend_risk') or {}
        risk_triggers = 0
        if float(frontend.get('cpi_increase_pct') or 0) >= float(risk_cfg.get('cpi_increase_pct') or 25):
            risk_triggers += 1
        if float(frontend.get('ctr_decrease_pct') or 0) >= float(risk_cfg.get('ctr_decrease_pct') or 20):
            risk_triggers += 1
        if float(frontend.get('cpm_increase_pct') or 0) >= float(risk_cfg.get('cpm_increase_pct') or 30):
            risk_triggers += 1
        frontend_risky = risk_triggers >= int(risk_cfg.get('min_trigger_count') or 2)

        action = 'observe'
        pct = 0
        reason = '数据尚需继续观察。'
        status_tag = 'observe'
        gate_warnings = set(data_quality_gate.warnings or []) if data_quality_gate is not None else set()
        if data_quality_gate is not None and (
            data_quality_gate.status == 'BLOCKED'
            or 'no_success_bind_events' in gate_warnings
            or 'dedupe_not_confirmed' in gate_warnings
        ):
            action, reason, status_tag = 'observe', 'TUGAO 真实入会数据未通过完整性检查，本轮只展示数据异常，不生成暂停、降预算或放量建议。', 'data_missing'
        elif item.data_quality.status != 'ok':
            action, reason, status_tag = 'manual_review', '数据质量不足，进入人工复核。', 'data_quality'
        elif cap is None:
            action, reason, status_tag = 'observe', '该市场暂无真实入会成本红线，仅展示相对排名和漏斗诊断，不给强动作。', 'no_cap'
        elif binds == 0 and maturity >= int(rules.get('maturity_day') or 2) and spend >= cap * float(rules.get('zero_conversion_spend_multiplier') or 3):
            action, reason, status_tag = 'pause', 'D+2 后消耗已达到国家红线 3 倍且真实入会为 0，建议暂停。', 'zero_stop'
        elif binds == 0 and maturity >= int(rules.get('under_delivery_review_day') or 4) and spend >= cap * float(rules.get('sample_insufficient_spend_multiplier') or 1):
            action, reason, status_tag = 'manual_review', 'D+4 仍未达到有效消耗门槛，标记为投放不足并人工复核。', 'under_delivery'
        elif binds == 0 and spend < cap * float(rules.get('sample_insufficient_spend_multiplier') or 1):
            action, reason, status_tag = 'observe', f'消耗 ${spend:.2f} 低于样本门槛 ${cap * float(rules.get("sample_insufficient_spend_multiplier") or 1):.2f}，不足以判断投放失败。', 'sample_insufficient'
        elif frontend_risky and cpa is not None and cpa <= cap:
            action, reason, status_tag = 'hold_scale', '真实入会成本达标，但安装成本、点击率或千次展示成本同步恶化，暂缓放量。', 'frontend_risk'
        elif binds >= int(rules.get('min_formal_winner_binds') or 10) and cpa is not None and cpa <= cap:
            if binds >= int(rules.get('high_confidence_binds') or 20) and cpa <= cap * 0.85:
                pct = 15
            elif binds >= 18 and cpa <= cap * 0.75:
                pct = 20
            else:
                pct = 10
            action, reason, status_tag = 'scale_up', f'真实入会 {binds}，真实入会成本 ${cpa:.2f} 低于国家红线 ${cap:.2f}。', 'winner'
        elif binds > 0 and binds < int(rules.get('min_formal_winner_binds') or 10) and cpa is not None and cpa <= cap:
            action, reason, status_tag = 'observe', '低样本但真实入会成本达标，标记为潜力广告并继续观察。', 'potential_winner'
        elif cpa is not None and binds < int(rules.get('min_formal_winner_binds') or 10) and spend < cap * float(rules.get('zero_conversion_spend_multiplier') or 3):
            action, reason, status_tag = 'observe', f'真实入会样本 {binds} 个且消耗 ${spend:.2f} 未达到止损线 ${cap * float(rules.get("zero_conversion_spend_multiplier") or 3):.2f}，暂不按 CPA 异常给强动作。', 'sample_insufficient'
        elif cpa is not None and cpa >= cap * 1.40:
            action, reason, status_tag = 'pause', f'真实入会成本 ${cpa:.2f} 严重超过国家红线 ${cap:.2f}。', 'severe_over_cap'
        elif cpa is not None and cpa >= cap * 1.15:
            action, pct, reason, status_tag = 'reduce_budget', 15, f'真实入会成本 ${cpa:.2f} 中度超过国家红线 ${cap:.2f}。', 'over_cap'
        elif cpa is not None and cpa > cap:
            action, reason, status_tag = 'observe', f'真实入会成本 ${cpa:.2f} 轻微超过国家红线 ${cap:.2f}，先继续观察。', 'slight_over_cap'

        if item.budget_mode.mode == 'mixed_change':
            action, pct, reason, status_tag = 'observe', 0, '检测到加预算同时更换素材，本阶段只展示效果，不归因建议有效性。', 'mixed_change'
        if status_tag == 'severe_over_cap' and item.campaign and '结构优化' in item.campaign:
            reason = '广告系列整体超标，但内部存在优质广告组；建议保留赢家、暂停拖累项，不暂停整个系列。'
            status_tag = 'structure_optimization'

        confidence = 'high' if binds >= int(rules.get('high_confidence_binds') or 20) else ('medium' if binds >= 10 else 'low')
        if action in {'pause', 'reduce_budget'} and binds >= 10:
            confidence = 'medium'
        if action == 'manual_review':
            confidence = 'low'
        evidence = RecommendationEvidence(
            spend=item.spend,
            installs=item.installs,
            real_bind_count=binds,
            real_bind_cpa=item.real_bind_cpa,
            cpi=item.cpi,
            ctr=item.ctr,
            cpm=item.cpm,
            country_cap=cap,
            data_window=window,
            data_quality=item.data_quality,
            budget_mode=item.budget_mode,
            frontend_trend=item.frontend_trend,
            funnel_metrics=self.funnel_metrics_payload(item),
            evidence_points=[
                f'真实入会 {binds}',
                '-' if item.real_bind_cpa is None else f'真实入会成本 ${item.real_bind_cpa:.2f}',
            ],
        )
        name = item.ad or item.ad_group or item.campaign or item.object_id
        diagnosis_type = _diagnosis_type_from_status(status_tag)
        action_type = _action_type_from_diagnosis(action, diagnosis_type)
        gate = self.action_gate(
            item,
            diagnosis_type=diagnosis_type,
            action_type=action_type,
            allow_generate_creative=status_tag in LEGACY_GENERATIVE_STATUS_TAGS,
            allow_pause=action in {'pause', 'reduce_budget'},
            allow_scale=action == 'scale_up',
            allow_reduce_budget=action == 'reduce_budget',
        )
        return Recommendation(
            recommendation_id=_stable_id(item.observation_identity or item.object_id, action, pct, RECOMMENDATION_RULE_VERSION),
            object_id=item.object_id,
            object_level=item.object_level,
            object_name=name,
            country=item.country,
            project=item.project,
            primary_action=action,
            primary_action_zh=ZH_LABELS.get(action, action),
            adjustment_pct=pct,
            reason_zh=reason,
            confidence=confidence,
            confidence_zh={'high': '高', 'medium': '中', 'low': '低'}.get(confidence, confidence),
            status_tag=status_tag,
            evidence=evidence,
            created_at_utc=datetime.now(timezone.utc).isoformat(),
            diagnosis_type=diagnosis_type,
            diagnosis_type_zh=ZH_LABELS.get(diagnosis_type, diagnosis_type),
            action_type=action_type,
            action_type_zh=ZH_LABELS.get(action_type, action_type),
            primary_layer='legacy_final_cpa',
            maturity_status='fully_matured' if binds >= int(rules.get('min_formal_winner_binds') or 10) else 'legacy',
            data_quality_status=item.data_quality.status,
            allow_pause=gate['allow_pause'],
            allow_scale=gate['allow_scale'],
            allow_generate_creative=gate['allow_generate_creative'],
            needs_data=['进入 IM', '用户行为型有效 IM', '链接点击/注册/bind', 'Linky/bind/CRM 漏斗'],
            action_gate=gate,
            decision_context={
                'evidence': {
                    'points': list(evidence.evidence_points),
                    'data_window': dict(window),
                    'rule_version': RECOMMENDATION_RULE_VERSION,
                },
                'risk': {
                    'blocked': bool(gate['blocked_reasons']),
                    'reasons': list(gate['blocked_reasons']),
                },
                'confidence': confidence,
                'recommended_action': action,
                'diagnosis_type': diagnosis_type,
            },
            observation_identity=item.observation_identity,
        )


def build_daily_report_from_dashboard_snapshot(
    snapshot: Dict[str, Any],
    *,
    report_date: Optional[str] = None,
    data_mode: str = 'fixture',
    provider: Optional[RealConversionProvider] = None,
    project: Optional[str] = None,
    country: Optional[str] = None,
    window_days: int = 1,
) -> DailyAdReportV1:
    target_date, window_start, window_end = london_report_window(report_date)
    normalized_window_days = min(max(int(window_days or 1), 1), 31)
    if normalized_window_days > 1:
        start_date = target_date - timedelta(days=normalized_window_days - 1)
        _, window_start, _ = london_report_window(start_date.isoformat())
    if provider is None:
        provider = FixtureRealConversionProvider()
    events = provider.get_bind_events(window_start, window_end, project=project, country=country)
    dedupe_confirmed = isinstance(provider, TugaoRealConversionProvider)
    real_bind_metric = summarize_real_bind_metric_contract(events, dedupe_confirmed=dedupe_confirmed)
    objects = adapt_dashboard_snapshot_to_report_objects(
        snapshot,
        events,
        target_date,
        allow_fixture_fallback=str(data_mode or '').strip().lower() == 'fixture',
    )
    if country:
        wanted = normalize_country_code(country)
        objects = [item for item in objects if item.country == wanted]
    fact_bind_count = sum(int(item.real_bind_count or 0) for item in objects)
    if real_bind_metric.final_real_bind_count <= 0 and fact_bind_count > 0:
        real_bind_metric = replace(
            real_bind_metric,
            attribution_version='tugao_funnel_fact_v1',
            is_dedupe_confirmed=True,
            bind_event_count=fact_bind_count,
            unique_bind_count=fact_bind_count,
            unique_customer_user_count=fact_bind_count,
            final_real_bind_count=fact_bind_count,
            missing_country_count=0,
            missing_project_count=0,
            missing_attribution_count=0,
        )
    data_quality_gate = evaluate_data_quality_gate(real_bind_metric)
    engine = AdDailyRecommendationEngine()
    window = {
        'timezone': REPORT_TZ_NAME,
        'start_utc': window_start.isoformat(),
        'end_utc': window_end.isoformat(),
        'report_date': target_date.isoformat(),
        'window_days': normalized_window_days,
    }
    recommendation_data_gate = data_quality_gate if str(data_mode or '').strip().lower() != 'fixture' else None
    actionable_objects = [
        item for item in objects
        if not _is_historical_settlement_placeholder(item)
        and _is_paid_ad_scoring_candidate(item)
    ]
    recommendations = sort_recommendations_by_intervention_priority(
        engine.recommend(item, window, data_quality_gate=recommendation_data_gate)
        for item in actionable_objects
    )
    summary_objects = [
        item for item in objects
        if _is_historical_settlement_placeholder(item)
        or _is_paid_ad_scoring_candidate(item)
    ]
    summary = build_report_summary(summary_objects, recommendations)
    summary.update({
        'data_quality_status': data_quality_gate.status,
        'data_quality_status_zh': data_quality_gate.status_zh,
        'dedupe_version': real_bind_metric.dedupe_version,
        'attribution_version': real_bind_metric.attribution_version,
        'real_bind_count_mode': real_bind_metric.real_bind_count_mode,
        'real_bind_count_mode_label_cn': real_bind_metric.real_bind_count_mode_label_cn,
        'is_dedupe_confirmed': real_bind_metric.is_dedupe_confirmed,
        'bind_event_count': real_bind_metric.bind_event_count,
        'unique_bind_count': real_bind_metric.unique_bind_count,
        'unique_customer_user_count': real_bind_metric.unique_customer_user_count,
        'final_real_bind_count': real_bind_metric.final_real_bind_count,
        'report_window_days': normalized_window_days,
        'report_window_label': (
            f'{(target_date - timedelta(days=normalized_window_days - 1)).isoformat()} 至 {target_date.isoformat()}'
            if normalized_window_days > 1
            else target_date.isoformat()
        ),
    })
    report_id = _stable_id(REPORT_SCHEMA_VERSION, target_date.isoformat(), data_mode, normalized_window_days, project or '', country or '')
    report = DailyAdReportV1(
        report_id=f'ad_daily_{target_date.strftime("%Y%m%d")}_{report_id[:8]}',
        snapshot_version=REPORT_SCHEMA_VERSION,
        rule_version=RECOMMENDATION_RULE_VERSION,
        report_date=target_date.isoformat(),
        window_start_utc=window_start.isoformat(),
        window_end_utc=window_end.isoformat(),
        window_timezone=REPORT_TZ_NAME,
        data_mode='模拟' if data_mode == 'fixture' else str(data_mode),
        provider='fixture' if isinstance(provider, FixtureRealConversionProvider) else provider.__class__.__name__,
        generated_at_utc=datetime.now(timezone.utc).isoformat(),
        summary=summary,
        real_bind_metric=real_bind_metric,
        data_quality_gate=data_quality_gate,
        ad_objects=objects,
        recommendations=sort_recommendations_by_intervention_priority(recommendations),
        creative_test_plan=build_creative_test_plan(objects, recommendations),
        creative_insights={},
        review_skeleton=build_review_skeleton(recommendations, data_quality_gate=data_quality_gate),
    )
    return replace(report, creative_insights=build_creative_intelligence_payload(report))


def _is_historical_settlement_placeholder(item: AdObjectMetrics) -> bool:
    account = str(item.account_id or '').strip().lower()
    if account in {value.lower() for value in HISTORICAL_SETTLEMENT_ACCOUNT_LABELS}:
        return True
    return all(
        str(value or '').strip() == HISTORICAL_SETTLEMENT_PLACEHOLDER
        for value in (item.campaign, item.ad_group, item.ad)
    )


def _is_paid_ad_scoring_candidate(item: AdObjectMetrics) -> bool:
    """Keep raw facts intact while limiting recommendations to attributable paid ads."""
    source_type = str(item.source_type or '').strip().lower()
    if any(marker in source_type for marker in ('自然', 'organic')):
        return False

    account = str(item.account_id or '').strip()
    if not account or account.lower() in {
        'unknown', 'unknown_account', 'internal', '未归属账户', '未归属广告账户',
    }:
        return False

    ad_name = str(item.ad or '').strip()
    if not ad_name or ad_name.lower() in {'unknown', 'unnamed', '未命名', '自然量'}:
        return False
    return True


def build_report_summary(objects: List[AdObjectMetrics], recommendations: List[Recommendation]) -> Dict[str, Any]:
    total_spend = sum(float(item.spend or 0.0) for item in objects)
    total_binds = sum(int(item.real_bind_count or 0) for item in objects)
    action_counts: Dict[str, int] = {}
    for reco in recommendations:
        action_counts[reco.primary_action] = action_counts.get(reco.primary_action, 0) + 1
    return {
        'real_bind_count': total_binds,
        'real_bind_cpa': round(total_spend / total_binds, 4) if total_binds else None,
        'scale_up_count': action_counts.get('scale_up', 0),
        'reduce_budget_count': action_counts.get('reduce_budget', 0),
        'pause_count': action_counts.get('pause', 0),
        'observe_count': action_counts.get('observe', 0),
        'manual_review_count': action_counts.get('manual_review', 0),
        'hold_scale_count': action_counts.get('hold_scale', 0),
        'total_spend': round(total_spend, 4),
        'object_count': len(objects),
        'recommendation_count': len(recommendations),
        'updated_at_utc': datetime.now(timezone.utc).isoformat(),
    }


def sort_recommendations_by_intervention_priority(recommendations: Iterable[Recommendation]) -> List[Recommendation]:
    diagnosis_priority = {
        'data_anomaly': 0,
        'front_funnel_weak': 1,
        'low_quality_traffic': 2,
        'creative_effective_post_im_failed': 3,
        'im_handoff_issue': 3,
        'cs_response_issue': 3,
        'linky_crm_issue': 3,
        'creative_fatigue': 5,
        'audience_mismatch': 5,
        'continue_observe': 6,
        'scale_opportunity': 4,
        'sample_insufficient': 7,
    }

    action_priority = {
        'pause': 0,
        'reduce_budget': 1,
        'manual_review': 2,
        'scale_up': 4,
        'hold_scale': 5,
        'observe': 6,
    }

    return sorted(
        list(recommendations or []),
        key=lambda reco: (
            7 if str(reco.status_tag or '') == 'sample_insufficient' else diagnosis_priority.get(str(getattr(reco, 'diagnosis_type', '') or reco.status_tag or ''), 6),
            action_priority.get(str(reco.primary_action or 'observe'), 9),
            -float(reco.evidence.spend or 0.0),
            -(float(reco.evidence.real_bind_cpa) if reco.evidence.real_bind_cpa is not None else 0.0),
            str(reco.object_name or ''),
        ),
    )


def build_creative_test_plan(objects: List[AdObjectMetrics], recommendations: List[Recommendation]) -> List[Dict[str, Any]]:
    by_id = {reco.object_id: reco for reco in recommendations}
    plans: List[Dict[str, Any]] = []
    for item in objects[:20]:
        reco = by_id.get(item.object_id)
        if not reco:
            continue
        if not bool(getattr(reco, 'allow_generate_creative', False)):
            continue
        if reco.status_tag in {'winner', 'potential_winner'} or reco.diagnosis_type == 'scale_opportunity':
            direction = '赢家延展'
            analysis = '保留当前核心卖点，扩展同语种素材角度。'
        elif reco.diagnosis_type in {'front_funnel_weak', 'low_quality_traffic'} or reco.status_tag in {'frontend_risk', 'over_cap', 'severe_over_cap'}:
            direction = '漏斗修复'
            analysis = '优先修正素材前链路、用户行为型有效 IM 和链接点击/注册/bind。'
        else:
            direction = '受控探索'
            analysis = '低风险扩展测试，不机械生成固定文案。'
        plans.append({
            'object_id': item.object_id,
            'country': item.country,
            'campaign': item.campaign,
            'creative_tags': [direction],
            'diagnosis_zh': analysis,
            'localized_copy': {
                'ID': {'language': 'id', 'text': 'Uji variasi manfaat utama.', 'zh': '测试核心利益点的不同表达。'},
                'BR': {'language': 'pt-BR', 'text': 'Teste novas provas sociais.', 'zh': '测试新的社会证明表达。'},
                'RECOMPA': {'language': 'es', 'text': 'Probar nuevos enfoques controlados.', 'zh': '测试受控的新角度。'},
            }.get(item.country, {'language': 'local', 'text': '', 'zh': '等待本地化文案。'}),
        })
    return plans


def build_review_skeleton(
    recommendations: List[Recommendation],
    *,
    data_quality_gate: Optional[DataQualityGateResult] = None,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for reco in recommendations:
        production_gate = evaluate_production_recommendation_gate(
            reco,
            mode='shadow',
            data_quality_gate=data_quality_gate,
        )
        rows.append({
            'recommendation_id': reco.recommendation_id,
            'execution_status': '未执行',
            'shadow_rule_version': RECOMMENDATION_RULE_VERSION,
            'diagnosis_type': reco.diagnosis_type,
            'diagnosis_type_zh': reco.diagnosis_type_zh,
            'action_type': reco.action_type,
            'action_type_zh': reco.action_type_zh,
            'primary_layer': reco.primary_layer,
            'maturity_status': reco.maturity_status,
            'allow_pause': reco.allow_pause,
            'allow_scale': reco.allow_scale,
            'creative_scale_candidate': reco.creative_scale_candidate,
            'business_scale_allowed': reco.business_scale_allowed,
            'allow_generate_creative': reco.allow_generate_creative,
            'evidence_points': list(reco.evidence.evidence_points or []),
            'needs_data': list(reco.needs_data or []),
            'production_gate': asdict(production_gate),
            'data_quality_gate': asdict(data_quality_gate) if data_quality_gate else {},
            'review_days': ['D+1', 'D+2', 'D+3'],
            'primary_baseline': '执行前连续 3 天均值',
            'secondary_baseline': '上周相同星期',
            'effective_condition': '干净执行且前后窗口真实入会均不少于 10，放量后真实入会增长不少于 10%，真实入会成本未超过国家红线。',
        })
    return rows


def report_to_dict(report: DailyAdReportV1) -> Dict[str, Any]:
    return asdict(report)


def report_from_dict(payload: Dict[str, Any]) -> DailyAdReportV1:
    ad_objects = [
        AdObjectMetrics(
            **{
                **dict(item),
                'data_quality': DataQualityStatus(**dict(item.get('data_quality') or {})),
                'budget_mode': BudgetMode(**dict(item.get('budget_mode') or {})),
            }
        )
        for item in payload.get('ad_objects') or []
    ]
    recommendations = []
    for item in payload.get('recommendations') or []:
        raw_item = dict(item or {})
        status_tag = str(raw_item.get('status_tag') or '').strip()
        primary_action = str(raw_item.get('primary_action') or 'observe').strip()
        if not raw_item.get('diagnosis_type') or (
            raw_item.get('diagnosis_type') == 'continue_observe'
            and status_tag
            and status_tag not in {'observe', 'slight_over_cap', 'potential_winner', 'no_cap', 'mixed_change'}
        ):
            raw_item['diagnosis_type'] = _diagnosis_type_from_status(status_tag)
        raw_item['diagnosis_type_zh'] = raw_item.get('diagnosis_type_zh') or ZH_LABELS.get(str(raw_item.get('diagnosis_type') or ''), str(raw_item.get('diagnosis_type') or ''))
        if not raw_item.get('action_type') or (
            raw_item.get('action_type') == 'observe'
            and str(raw_item.get('diagnosis_type') or '') in {'front_funnel_weak', 'creative_fatigue', 'scale_opportunity'}
        ):
            raw_item['action_type'] = _action_type_from_diagnosis(primary_action, str(raw_item.get('diagnosis_type') or ''))
        raw_item['action_type_zh'] = raw_item.get('action_type_zh') or ZH_LABELS.get(str(raw_item.get('action_type') or ''), str(raw_item.get('action_type') or ''))
        raw_item.setdefault('allow_pause', primary_action in {'pause', 'reduce_budget'})
        raw_item.setdefault('allow_scale', primary_action == 'scale_up')
        raw_item.setdefault('creative_scale_candidate', raw_item.get('diagnosis_type') == 'creative_scale_candidate')
        raw_item.setdefault('business_scale_allowed', bool(raw_item.get('allow_scale')))
        raw_item.setdefault('allow_generate_creative', status_tag in LEGACY_GENERATIVE_STATUS_TAGS)
        raw_item.setdefault('needs_data', ['用户报名人数', '用户行为型有效 IM', '链接点击', 'Linky/bind/CRM 漏斗'])
        raw_item.setdefault('creative_diagnosis', {})
        raw_item.setdefault('post_im_diagnosis', {})
        raw_item.setdefault('business_diagnosis', {})
        raw_item.setdefault('data_origin', 'LEGACY')
        evidence = dict(raw_item.get('evidence') or {})
        recommendations.append(Recommendation(
            **{
                **raw_item,
                'evidence': RecommendationEvidence(
                    **{
                        **evidence,
                        'data_quality': DataQualityStatus(**dict(evidence.get('data_quality') or {})),
                        'budget_mode': BudgetMode(**dict(evidence.get('budget_mode') or {})),
                    }
                ),
            }
        ))
    return DailyAdReportV1(
        report_id=str(payload.get('report_id') or ''),
        snapshot_version=str(payload.get('snapshot_version') or REPORT_SCHEMA_VERSION),
        rule_version=str(payload.get('rule_version') or RECOMMENDATION_RULE_VERSION),
        report_date=str(payload.get('report_date') or ''),
        window_start_utc=str(payload.get('window_start_utc') or ''),
        window_end_utc=str(payload.get('window_end_utc') or ''),
        window_timezone=str(payload.get('window_timezone') or 'Europe/London'),
        data_mode=str(payload.get('data_mode') or '模拟'),
        provider=str(payload.get('provider') or 'fixture'),
        generated_at_utc=str(payload.get('generated_at_utc') or ''),
        summary=dict(payload.get('summary') or {}),
        real_bind_metric=RealBindMetricContract(**dict(payload.get('real_bind_metric') or {
            'dedupe_version': TUGAO_REAL_BIND_DEDUPE_VERSION,
            'attribution_version': TUGAO_REAL_BIND_ATTRIBUTION_VERSION,
            'real_bind_count_mode': REAL_BIND_COUNT_MODE,
            'real_bind_count_mode_label_cn': REAL_BIND_COUNT_MODE_LABEL_CN,
            'is_dedupe_confirmed': False,
            'bind_event_count': 0,
            'unique_bind_count': 0,
            'unique_customer_user_count': 0,
            'final_real_bind_count': 0,
            'has_wa_success_count': 0,
            'no_wa_success_count': 0,
        })),
        data_quality_gate=DataQualityGateResult(**dict(payload.get('data_quality_gate') or {
            'status': 'WARNING',
            'status_zh': '预警',
            'reasons': [],
            'warnings': ['legacy_report_missing_gate'],
            'checked_at_utc': '',
        })),
        ad_objects=ad_objects,
        recommendations=recommendations,
        creative_test_plan=list(payload.get('creative_test_plan') or []),
        creative_insights=dict(payload.get('creative_insights') or {}),
        review_skeleton=list(payload.get('review_skeleton') or []),
        labels=dict(payload.get('labels') or ZH_LABELS),
        simulation_notice=str(payload.get('simulation_notice') or '当前为模拟真实绑定数据，仅用于系统验证，不代表生产投放结论。'),
    )


def ensure_ad_daily_report_tables(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS ad_daily_report (
            report_id TEXT PRIMARY KEY,
            report_date TEXT NOT NULL,
            data_mode TEXT NOT NULL,
            snapshot_version TEXT NOT NULL,
            rule_version TEXT NOT NULL,
            window_start_utc TEXT NOT NULL,
            window_end_utc TEXT NOT NULL,
            generated_at_utc TEXT NOT NULL,
            payload_json TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS ad_recommendation (
            recommendation_id TEXT PRIMARY KEY,
            report_id TEXT NOT NULL,
            object_id TEXT NOT NULL,
            primary_action TEXT NOT NULL,
            primary_action_zh TEXT NOT NULL,
            confidence TEXT NOT NULL,
            status_tag TEXT NOT NULL,
            decision_context_json TEXT NOT NULL DEFAULT '{}',
            data_origin TEXT NOT NULL DEFAULT 'LEGACY',
            payload_json TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS ad_recommendation_evidence (
            recommendation_id TEXT PRIMARY KEY,
            report_id TEXT NOT NULL,
            evidence_json TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS ad_recommendation_review (
            recommendation_id TEXT PRIMARY KEY,
            report_id TEXT NOT NULL,
            review_json TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS ad_production_recommendation_gate (
            recommendation_id TEXT PRIMARY KEY,
            report_id TEXT NOT NULL,
            gate_status TEXT NOT NULL,
            publishable INTEGER NOT NULL,
            payload_json TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS ad_creative_test_plan (
            plan_id TEXT PRIMARY KEY,
            report_id TEXT NOT NULL,
            payload_json TEXT NOT NULL
        );
        """
    )
    columns = {row[1] for row in conn.execute('PRAGMA table_info(ad_recommendation)').fetchall()}
    if 'decision_context_json' not in columns:
        conn.execute("ALTER TABLE ad_recommendation ADD COLUMN decision_context_json TEXT NOT NULL DEFAULT '{}'")
    if 'data_origin' not in columns:
        conn.execute("ALTER TABLE ad_recommendation ADD COLUMN data_origin TEXT NOT NULL DEFAULT 'LEGACY'")


def persist_daily_report(conn: sqlite3.Connection, report: DailyAdReportV1) -> None:
    ensure_ad_daily_report_tables(conn)
    payload = report_to_dict(report)
    conn.execute(
        """
        INSERT OR REPLACE INTO ad_daily_report
        (report_id, report_date, data_mode, snapshot_version, rule_version, window_start_utc, window_end_utc, generated_at_utc, payload_json)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            report.report_id,
            report.report_date,
            report.data_mode,
            report.snapshot_version,
            report.rule_version,
            report.window_start_utc,
            report.window_end_utc,
            report.generated_at_utc,
            json.dumps(payload, ensure_ascii=False, sort_keys=True),
        ),
    )
    for reco in report.recommendations:
        reco_payload = asdict(reco)
        conn.execute(
            """
            INSERT OR REPLACE INTO ad_recommendation
            (recommendation_id, report_id, object_id, primary_action, primary_action_zh,
             confidence, status_tag, decision_context_json, data_origin, payload_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'NATIVE_V2', ?)
            """,
            (
                reco.recommendation_id,
                report.report_id,
                reco.object_id,
                reco.primary_action,
                reco.primary_action_zh,
                reco.confidence,
                reco.status_tag,
                json.dumps(reco.decision_context, ensure_ascii=False, sort_keys=True),
                json.dumps(reco_payload, ensure_ascii=False, sort_keys=True),
            ),
        )
        conn.execute(
            "INSERT OR REPLACE INTO ad_recommendation_evidence (recommendation_id, report_id, evidence_json) VALUES (?, ?, ?)",
            (reco.recommendation_id, report.report_id, json.dumps(asdict(reco.evidence), ensure_ascii=False, sort_keys=True)),
        )
    for review in report.review_skeleton:
        conn.execute(
            "INSERT OR REPLACE INTO ad_recommendation_review (recommendation_id, report_id, review_json) VALUES (?, ?, ?)",
            (review['recommendation_id'], report.report_id, json.dumps(review, ensure_ascii=False, sort_keys=True)),
        )
        production_gate = dict(review.get('production_gate') or {})
        if production_gate:
            conn.execute(
                """
                INSERT OR REPLACE INTO ad_production_recommendation_gate
                (recommendation_id, report_id, gate_status, publishable, payload_json)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    review['recommendation_id'],
                    report.report_id,
                    str(production_gate.get('gate_status') or 'blocked'),
                    1 if production_gate.get('publishable') else 0,
                    json.dumps(production_gate, ensure_ascii=False, sort_keys=True),
                ),
            )
    for index, plan in enumerate(report.creative_test_plan):
        conn.execute(
            "INSERT OR REPLACE INTO ad_creative_test_plan (plan_id, report_id, payload_json) VALUES (?, ?, ?)",
            (_stable_id(report.report_id, index, plan.get('object_id')), report.report_id, json.dumps(plan, ensure_ascii=False, sort_keys=True)),
        )
    from app.growth.ad_observation_materializer import materialize_observation_tasks

    materialize_observation_tasks(conn, payload)
    conn.commit()


def load_persisted_daily_report(
    conn: sqlite3.Connection,
    *,
    report_date: Optional[str] = None,
    data_mode: str = 'real',
) -> Optional[Dict[str, Any]]:
    ensure_ad_daily_report_tables(conn)
    normalized_mode = str(data_mode or 'real').strip().lower()
    stored_modes = ['模拟', 'fixture'] if normalized_mode == 'fixture' else [normalized_mode]
    params: List[Any] = []
    where = []
    if report_date:
        where.append('report_date = ?')
        params.append(str(report_date).strip())
    if stored_modes:
        where.append(f"data_mode IN ({','.join('?' for _ in stored_modes)})")
        params.extend(stored_modes)
    query = "SELECT payload_json FROM ad_daily_report"
    if where:
        query += " WHERE " + " AND ".join(where)
    query += " ORDER BY report_date DESC, generated_at_utc DESC LIMIT 1"
    row = conn.execute(query, params).fetchone()
    if not row:
        return None
    payload = json.loads(row['payload_json'] if isinstance(row, sqlite3.Row) else row[0])
    if isinstance(payload, dict):
        try:
            payload = report_to_dict(report_from_dict(payload))
        except Exception:
            payload = dict(payload)
        recommendation_ids = [
            str(item.get('recommendation_id') or '')
            for item in payload.get('recommendations') or []
            if isinstance(item, dict) and item.get('recommendation_id')
        ]
        if recommendation_ids:
            placeholders = ','.join('?' for _ in recommendation_ids)
            origin_rows = conn.execute(
                f"SELECT recommendation_id, data_origin FROM ad_recommendation WHERE recommendation_id IN ({placeholders})",
                recommendation_ids,
            ).fetchall()
            origins = {
                str(origin['recommendation_id']): str(origin['data_origin'] or 'LEGACY')
                for origin in origin_rows
            }
            for item in payload.get('recommendations') or []:
                if isinstance(item, dict):
                    item['data_origin'] = origins.get(str(item.get('recommendation_id') or ''), 'LEGACY')
        payload['served_from_cache'] = True
        return payload
    return None


def recommendation_history_payload(conn: sqlite3.Connection, limit: int = 100) -> Dict[str, Any]:
    ensure_ad_daily_report_tables(conn)
    rows = conn.execute(
        """
        SELECT r.report_id, r.report_date, r.data_mode, r.rule_version, a.data_origin, a.payload_json
        FROM ad_recommendation a
        JOIN ad_daily_report r ON r.report_id = a.report_id
        ORDER BY r.generated_at_utc DESC, a.recommendation_id ASC
        LIMIT ?
        """,
        (max(1, min(int(limit or 100), 500)),),
    ).fetchall()
    items = []
    for row in rows:
        payload = json.loads(row['payload_json'])
        items.append({
            'report_id': row['report_id'],
            'report_date': row['report_date'],
            'data_mode': row['data_mode'],
            'rule_version': row['rule_version'],
            **payload,
            'data_origin': str(row['data_origin'] or 'LEGACY'),
        })
    return {'items': items, 'count': len(items)}


def recommendation_review_payload(conn: sqlite3.Connection, recommendation_id: str) -> Dict[str, Any]:
    ensure_ad_daily_report_tables(conn)
    row = conn.execute(
        """
        SELECT a.data_origin, a.payload_json AS recommendation_json, e.evidence_json, v.review_json
        FROM ad_recommendation a
        LEFT JOIN ad_recommendation_evidence e ON e.recommendation_id = a.recommendation_id
        LEFT JOIN ad_recommendation_review v ON v.recommendation_id = a.recommendation_id
        WHERE a.recommendation_id = ?
        """,
        (recommendation_id,),
    ).fetchone()
    if not row:
        return {'detail': 'recommendation_not_found'}
    recommendation = json.loads(row['recommendation_json'])
    recommendation['data_origin'] = str(row['data_origin'] or 'LEGACY')
    return {
        'recommendation': recommendation,
        'evidence': json.loads(row['evidence_json']) if row['evidence_json'] else {},
        'review': json.loads(row['review_json']) if row['review_json'] else {},
    }


def export_daily_report_xlsx(report: DailyAdReportV1) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = '仪表盘'
    header_fill = PatternFill('solid', fgColor='1F2937')
    header_font = Font(color='FFFFFF', bold=True)

    def write_table(sheet: Any, rows: List[List[Any]], start_row: int = 1) -> None:
        for r_index, row in enumerate(rows, start_row):
            for c_index, value in enumerate(row, 1):
                cell = sheet.cell(row=r_index, column=c_index, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if r_index == start_row:
                    cell.fill = header_fill
                    cell.font = header_font
        for col in range(1, (len(rows[0]) if rows else 1) + 1):
            sheet.column_dimensions[chr(64 + min(col, 26))].width = 18

    meta_rows = [
        ['字段', '值'],
        ['报告编号', report.report_id],
        ['数据快照版本', report.snapshot_version],
        ['规则版本', report.rule_version],
        ['去重版本', report.real_bind_metric.dedupe_version],
        ['归因版本', report.real_bind_metric.attribution_version],
        ['真实入会计数口径', report.real_bind_metric.real_bind_count_mode_label_cn],
        ['数据质量状态', report.data_quality_gate.status_zh],
        ['英国时间窗口', f"{report.report_date} Europe/London"],
        ['数据模式', report.data_mode],
        ['数据更新时间', report.generated_at_utc],
        ['模拟提示', report.simulation_notice],
        ['成功事件数', report.real_bind_metric.bind_event_count],
        ['唯一绑定数', report.real_bind_metric.unique_bind_count],
        ['唯一业务用户数', report.real_bind_metric.unique_customer_user_count],
        ['最终真实入会人数', report.real_bind_metric.final_real_bind_count],
        ['入会成功无 WA', report.real_bind_metric.no_wa_success_count],
        ['真实入会人数', report.summary.get('real_bind_count')],
        ['真实入会成本', report.summary.get('real_bind_cpa')],
    ]
    write_table(ws, meta_rows)

    ws2 = wb.create_sheet('账户汇总')
    write_table(ws2, [
        ['国家', '对象数', '真实入会人数', '消耗金额', '真实入会成本'],
        *[
            [
                country,
                len(items),
                sum(item.real_bind_count for item in items),
                round(sum(item.spend for item in items), 4),
                round(sum(item.spend for item in items) / sum(item.real_bind_count for item in items), 4) if sum(item.real_bind_count for item in items) else '',
            ]
            for country, items in _group_objects_by_country(report.ad_objects).items()
        ],
    ])

    ws3 = wb.create_sheet('广告表现')
    write_table(ws3, [
        ['国家', '广告账户', '广告系列', '广告组', '广告', '消耗金额', '展示次数', '点击次数', '点击率', '千次展示成本', '安装人数', '安装成本', '高价值用户', '高价值占比', '注册→用户报名率', '用户报名人数', '用户报名成本', '自动报名消息发送人数', '用户行为型有效IM', '有效IM口径', '用户行为型有效IM成本', '链接点击', 'Linky注册', 'bind成功', '用户报名→入会率', '真实入会人数', '真实入会成本'],
        *[
            [item.country, item.account_id, item.campaign, item.ad_group, item.ad, item.spend, item.impressions, item.clicks, item.ctr, item.cpm, item.installs, item.cpi, item.high_value_users, item.high_value_rate, item.registration_to_apply_rate, item.auto_apply_user_count, item.im_cost, item.auto_apply_message_users, item.user_engaged_im_users, item.user_engaged_im_metric_version, item.user_engaged_im_cost, item.link_click_users, item.linky_register_users, item.bind_success_users, item.im_to_join_rate, item.real_bind_count, item.real_bind_cpa]
            for item in report.ad_objects
        ],
    ])

    ws4 = wb.create_sheet('优化建议')
    write_table(ws4, [
        ['建议 ID', '国家', '对象', '主动作', '问题归因', '动作类型', '调整幅度', '主因', '建议置信度', '成熟度', '允许生成素材', '允许暂停', '允许放量', '素材放量候选', '经营允许放量', '证据', '缺失数据', '数据状态', '建议闸门', '阻断原因', '规则版本', '去重版本', '归因版本'],
        *[
            [
                reco.recommendation_id,
                reco.country,
                reco.object_name,
                reco.primary_action_zh,
                reco.diagnosis_type_zh,
                reco.action_type_zh,
                reco.adjustment_pct,
                reco.reason_zh,
                reco.confidence_zh,
                reco.maturity_status,
                '是' if reco.allow_generate_creative else '否',
                '是' if reco.allow_pause else '否',
                '是' if reco.allow_scale else '否',
                '是' if reco.creative_scale_candidate else '否',
                '是' if reco.business_scale_allowed else '否',
                '；'.join(reco.evidence.evidence_points or []),
                '；'.join(reco.needs_data or []),
                reco.status_tag,
                (next((row.get('production_gate') or {} for row in report.review_skeleton if row.get('recommendation_id') == reco.recommendation_id), {}) or {}).get('gate_status'),
                ', '.join((next((row.get('production_gate') or {} for row in report.review_skeleton if row.get('recommendation_id') == reco.recommendation_id), {}) or {}).get('reasons') or []),
                report.rule_version,
                report.real_bind_metric.dedupe_version,
                report.real_bind_metric.attribution_version,
            ]
            for reco in report.recommendations
        ],
    ])

    ws5 = wb.create_sheet('文案与素材测试')
    write_table(ws5, [
        ['国家', '广告系列', '素材方向', '漏斗诊断', '目标语言文案', '中文释义'],
        *[
            [plan.get('country'), plan.get('campaign'), ', '.join(plan.get('creative_tags') or []), plan.get('diagnosis_zh'), (plan.get('localized_copy') or {}).get('text'), (plan.get('localized_copy') or {}).get('zh')]
            for plan in report.creative_test_plan
        ],
    ])

    ws6 = wb.create_sheet('素材洞察')
    creative = report.creative_insights or {}
    creative_rows = creative.get('direction_performance') or []
    notice = '素材判断基于当前可见素材内容、广告表现和真实入会结果。若归因粒度为广告级或动态素材组合级，不能代表单个素材元素的独立贡献。'
    write_table(ws6, [
        ['说明', notice],
        ['同步素材数', (creative.get('status') or {}).get('synced_asset_count', 0)],
        ['可分析素材数', (creative.get('status') or {}).get('analyzable_asset_count', 0)],
        ['动态素材组合数', (creative.get('status') or {}).get('dynamic_creative_count', 0)],
    ])
    start_row = 7
    write_table(ws6, [
        ['素材方向', '消耗', '点击率', '安装成本', 'AF 模型入会', '真实入会', '真实入会成本', '素材判断', '下一步建议', '归因粒度', '分析置信度'],
        *[
            [
                row.get('direction'),
                row.get('spend'),
                row.get('ctr'),
                row.get('cpi'),
                row.get('af_model_join_events'),
                row.get('tugao_real_bind_count'),
                row.get('real_bind_cpa'),
                row.get('judgment'),
                row.get('next_step'),
                row.get('attribution_grain'),
                row.get('confidence'),
            ]
            for row in creative_rows
        ],
    ], start_row=start_row)

    ws7 = wb.create_sheet('信息流广告图生成')
    generated_images = ((creative.get('image_generation') or {}).get('generated_images') or [])
    write_table(ws7, [
        ['字段', '值'],
        ['默认素材面', '信息流广告图'],
        ['默认规格', '1024x1024'],
        ['构图要求', '方图、满版、无白边'],
        ['生产边界', '只生成投放初稿，不自动发布、不改预算'],
    ])
    write_table(ws7, [
        ['图片ID', '国家', '项目', '品牌', '规格', '审核状态', '风险状态', '图片引用'],
        *[
            [
                image.get('image_id'),
                image.get('country'),
                image.get('project'),
                image.get('brand'),
                image.get('image_size') or '1024x1024',
                image.get('review_status'),
                image.get('risk_status'),
                image.get('image_ref'),
            ]
            for image in generated_images
        ],
    ], start_row=7)

    ws8 = wb.create_sheet('源数据与对账')
    write_table(ws8, [
        ['字段', '值'],
        ['报告窗口 UTC 开始', report.window_start_utc],
        ['报告窗口 UTC 结束', report.window_end_utc],
        ['Provider', report.provider],
        ['数据质量 Gate', report.data_quality_gate.status],
        ['Gate 原因', ', '.join(report.data_quality_gate.reasons)],
        ['Gate 预警', ', '.join(report.data_quality_gate.warnings)],
        ['去重已确认', '是' if report.real_bind_metric.is_dedupe_confirmed else '否'],
        ['对象数', len(report.ad_objects)],
        ['建议数', len(report.recommendations)],
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _group_objects_by_country(objects: List[AdObjectMetrics]) -> Dict[str, List[AdObjectMetrics]]:
    grouped: Dict[str, List[AdObjectMetrics]] = {}
    for item in objects:
        grouped.setdefault(item.country, []).append(item)
    return grouped
