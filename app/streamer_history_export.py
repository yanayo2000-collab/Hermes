from __future__ import annotations

import io
import sqlite3
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta, timezone
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Set, Tuple
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from app.timo_guild_identity import timo_guild_display_name


BJ = ZoneInfo('Asia/Shanghai')
SUPPORTED_APPS = {'linky', 'timo'}


def normalize_history_app(value: object) -> str:
    app_name = str(value or '').strip().lower()
    if app_name not in SUPPORTED_APPS:
        raise ValueError('unsupported_streamer_history_app')
    return app_name


def normalize_streamer_id(value: object) -> str:
    streamer_id = str(value or '').strip()
    if not streamer_id or not streamer_id.isdigit():
        raise ValueError('invalid_streamer_id')
    return streamer_id


def _row_dict(row: Optional[sqlite3.Row]) -> Optional[Dict[str, Any]]:
    return dict(row) if row is not None else None


def lookup_streamer_first_join(
    conn: sqlite3.Connection,
    *,
    app_name: str,
    guild_name: str,
    streamer_id: str,
) -> Optional[Dict[str, Any]]:
    app = normalize_history_app(app_name)
    guild = str(guild_name or '').strip()
    sid = normalize_streamer_id(streamer_id)
    if not guild:
        raise ValueError('guild_name_required')
    conn.row_factory = sqlite3.Row
    if app == 'linky':
        row = conn.execute(
            """
            SELECT guild_executor_key, guild_name, streamer_id, platform_user_id,
                   platform_character_id, nickname, registered_at_bj, updated_at
            FROM streamer_external_profiles
            WHERE app_name = 'linky' AND guild_name = ?
              AND (streamer_id = ? OR platform_user_id = ? OR platform_character_id = ?)
              AND registered_at_bj <> ''
            ORDER BY registered_at_bj ASC
            LIMIT 1
            """,
            (guild, sid, sid, sid),
        ).fetchone()
        result = _row_dict(row)
        if result:
            result['canonical_streamer_id'] = str(result.get('streamer_id') or sid)
    else:
        row = conn.execute(
            """
            SELECT guild_executor_key, guild_name, timo_id, nickname,
                   registered_at_bj, updated_at
            FROM timo_external_streamers
            WHERE guild_name = ? AND timo_id = ? AND registered_at_bj <> ''
            ORDER BY registered_at_bj ASC
            LIMIT 1
            """,
            (guild, sid),
        ).fetchone()
        result = _row_dict(row)
        if result:
            result['canonical_streamer_id'] = str(result.get('timo_id') or sid)
    if not result:
        return None
    result.update({
        'app_name': app,
        'requested_streamer_id': sid,
        'first_join_date': str(result.get('registered_at_bj') or '')[:10],
    })
    return result


def load_local_revenue_rows(
    conn: sqlite3.Connection,
    *,
    profile: Dict[str, Any],
    date_to: str,
) -> List[Dict[str, Any]]:
    app = normalize_history_app(profile.get('app_name'))
    guild_name = str(profile.get('guild_name') or '')
    join_date = str(profile.get('first_join_date') or '')
    canonical_id = str(profile.get('canonical_streamer_id') or '')
    conn.row_factory = sqlite3.Row
    if app == 'linky':
        rows = conn.execute(
            """
            SELECT stat_date_bj, streamer_id, nickname, total_income, video_income
            FROM streamer_external_revenue_daily
            WHERE app_name = 'linky' AND guild_name = ? AND streamer_id = ?
              AND stat_date_bj BETWEEN ? AND ?
            ORDER BY stat_date_bj DESC
            """,
            (guild_name, canonical_id, join_date, date_to),
        ).fetchall()
        return [
            _normalized_revenue_row(
                stat_date=str(row['stat_date_bj']),
                streamer_id=canonical_id,
                nickname=str(row['nickname'] or profile.get('nickname') or ''),
                total_income=row['total_income'],
                voice_call_income=row['video_income'],
                source='local',
            )
            for row in rows
        ]
    rows = conn.execute(
        """
        SELECT stat_date_bj, timo_id, nickname, total_income, call_income
        FROM timo_external_revenue_daily
        WHERE guild_name = ? AND timo_id = ?
          AND stat_date_bj BETWEEN ? AND ?
        ORDER BY stat_date_bj DESC
        """,
        (guild_name, canonical_id, join_date, date_to),
    ).fetchall()
    return [
        _normalized_revenue_row(
            stat_date=str(row['stat_date_bj']),
            streamer_id=canonical_id,
            nickname=str(row['nickname'] or profile.get('nickname') or ''),
            total_income=row['total_income'],
            voice_call_income=row['call_income'],
            source='local',
        )
        for row in rows
    ]


def load_covered_dates(
    conn: sqlite3.Connection,
    *,
    profile: Dict[str, Any],
    date_to: str,
) -> Set[str]:
    app = normalize_history_app(profile.get('app_name'))
    guild_name = str(profile.get('guild_name') or '')
    join_date = str(profile.get('first_join_date') or '')
    conn.row_factory = sqlite3.Row
    if app == 'linky':
        rows = conn.execute(
            """
            SELECT stat_date_bj
            FROM streamer_external_guild_revenue_daily
            WHERE app_name = 'linky' AND guild_name = ?
              AND stat_date_bj BETWEEN ? AND ?
            """,
            (guild_name, join_date, date_to),
        ).fetchall()
    else:
        rows = conn.execute(
            """
            SELECT DISTINCT stat_date_bj
            FROM timo_external_revenue_daily
            WHERE guild_name = ? AND stat_date_bj BETWEEN ? AND ?
            """,
            (guild_name, join_date, date_to),
        ).fetchall()
    return {str(row['stat_date_bj']) for row in rows}


def _number(value: object) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _normalized_revenue_row(
    *,
    stat_date: str,
    streamer_id: str,
    nickname: str,
    total_income: object,
    voice_call_income: object,
    source: str,
) -> Dict[str, Any]:
    total = _number(total_income)
    voice = _number(voice_call_income)
    return {
        'stat_date': str(stat_date),
        'streamer_id': str(streamer_id),
        'nickname': str(nickname or ''),
        'total_income': total,
        'one_to_one_income': max(total - voice, 0.0),
        'voice_call_income': voice,
        'source': source,
    }


def normalize_timo_revenue_export_row(
    *,
    profile: Dict[str, Any],
    stat_date: str,
    row: Dict[str, Any],
) -> Dict[str, Any]:
    return _normalized_revenue_row(
        stat_date=stat_date,
        streamer_id=str(profile.get('requested_streamer_id') or profile.get('canonical_streamer_id') or ''),
        nickname=str(row.get('nick_name') or row.get('nickname') or profile.get('nickname') or ''),
        total_income=row.get('total_income'),
        voice_call_income=row.get('call_income'),
        source='live',
    )


def uncovered_dates(*, date_from: str, date_to: str, covered_dates: Iterable[str]) -> List[str]:
    start = date.fromisoformat(date_from)
    end = date.fromisoformat(date_to)
    if start > end:
        return []
    covered = {str(value) for value in covered_dates}
    return [
        (start + timedelta(days=offset)).isoformat()
        for offset in range((end - start).days + 1)
        if (start + timedelta(days=offset)).isoformat() not in covered
    ]


def fetch_linky_streamer_profile(
    *,
    executor: Dict[str, Any],
    streamer_id: str,
    fetcher: Optional[Callable[[Dict[str, Any], str, Dict[str, Any]], Dict[str, Any]]] = None,
) -> Optional[Dict[str, Any]]:
    sid = normalize_streamer_id(streamer_id)
    if fetcher is None:
        from app.streamer_external_sync import _linky_signed_get

        fetcher = _linky_signed_get
    payload = fetcher(executor, '/api/guild/search_anchors', {
        'id': sid,
        'page': 1,
        'page_size': 100,
    })
    rows = [dict(row) for row in (payload.get('items') or []) if isinstance(row, dict)]
    match = next(
        (
            row for row in rows
            if sid in {str(row.get('sid') or '').strip(), str(row.get('user_id') or '').strip()}
        ),
        None,
    )
    if not match:
        return None
    from app.streamer_external_sync import _bj_iso_from_epoch

    registered_at_bj = _bj_iso_from_epoch(match.get('created_at'))
    if not registered_at_bj:
        raise RuntimeError('linky_join_time_missing')
    canonical_sid = str(match.get('sid') or sid).strip()
    return {
        'app_name': 'linky',
        'guild_name': str(executor.get('guild_name') or ''),
        'streamer_id': canonical_sid,
        'canonical_streamer_id': canonical_sid,
        'requested_streamer_id': sid,
        'platform_user_id': str(match.get('user_id') or ''),
        'platform_character_id': canonical_sid,
        'nickname': str(match.get('nick_name') or ''),
        'registered_at_bj': registered_at_bj,
        'first_join_date': registered_at_bj[:10],
        'updated_at': datetime.now(timezone.utc).isoformat(),
        'lookup_source': 'linky_live',
    }


def _date_chunks(start: date, end: date, days: int = 31) -> List[Tuple[date, date]]:
    chunks: List[Tuple[date, date]] = []
    cursor = start
    while cursor <= end:
        chunk_end = min(end, cursor + timedelta(days=max(1, days) - 1))
        chunks.append((cursor, chunk_end))
        cursor = chunk_end + timedelta(days=1)
    return chunks


def _parse_linky_date(value: object) -> str:
    text = str(value or '').strip().replace('/', '-')
    return date.fromisoformat(text[:10]).isoformat()


def fetch_linky_streamer_history(
    *,
    executor: Dict[str, Any],
    streamer_id: str,
    date_from: str,
    date_to: str,
    fetcher: Optional[Callable[[Dict[str, Any], str, Dict[str, Any]], Dict[str, Any]]] = None,
    max_workers: int = 2,
) -> List[Dict[str, Any]]:
    sid = normalize_streamer_id(streamer_id)
    start = date.fromisoformat(date_from)
    end = date.fromisoformat(date_to)
    if start > end:
        return []
    if fetcher is None:
        from app.streamer_external_sync import _linky_signed_get

        fetcher = _linky_signed_get

    def fetch_chunk(chunk: Tuple[date, date]) -> Sequence[Dict[str, Any]]:
        chunk_start, chunk_end = chunk
        payload = fetcher(executor, '/api/guild/streamer_stat', {
            'begin': int(chunk_start.strftime('%Y%m%d')),
            'end': int(chunk_end.strftime('%Y%m%d')),
            'page_num': 1,
            'page_size': 100,
            'type': 0,
            'sid': sid,
        })
        return [dict(item) for item in (payload.get('items') or []) if isinstance(item, dict)]

    raw_rows: List[Dict[str, Any]] = []
    chunks = _date_chunks(start, end)
    with ThreadPoolExecutor(max_workers=max(1, min(int(max_workers or 1), 2))) as pool:
        futures = [pool.submit(fetch_chunk, chunk) for chunk in chunks]
        for future in as_completed(futures):
            raw_rows.extend(future.result())

    by_date: Dict[str, Dict[str, Any]] = {}
    for raw in raw_rows:
        raw_sid = str(raw.get('sid') or '').strip()
        if raw_sid != sid:
            continue
        stat_date = _parse_linky_date(raw.get('date'))
        by_date[stat_date] = _normalized_revenue_row(
            stat_date=stat_date,
            streamer_id=sid,
            nickname=str(raw.get('nickname') or ''),
            total_income=raw.get('total_earns'),
            voice_call_income=raw.get('voice_call_earns'),
            source='live',
        )
    return [by_date[key] for key in sorted(by_date, reverse=True)]


def merge_revenue_calendar(
    *,
    profile: Dict[str, Any],
    date_to: str,
    local_rows: Iterable[Dict[str, Any]],
    live_rows: Iterable[Dict[str, Any]] = (),
    covered_dates: Iterable[str] = (),
    live_full_range: bool = False,
) -> List[Dict[str, Any]]:
    start = date.fromisoformat(str(profile.get('first_join_date') or ''))
    end = date.fromisoformat(date_to)
    row_by_date = {str(row.get('stat_date')): dict(row) for row in local_rows}
    row_by_date.update({str(row.get('stat_date')): dict(row) for row in live_rows})
    covered = {str(value) for value in covered_dates}
    if live_full_range:
        covered.update((start + timedelta(days=offset)).isoformat() for offset in range((end - start).days + 1))
    result: List[Dict[str, Any]] = []
    cursor = end
    while cursor >= start:
        day_text = cursor.isoformat()
        row = row_by_date.get(day_text) or _normalized_revenue_row(
            stat_date=day_text,
            streamer_id=str(profile.get('requested_streamer_id') or profile.get('canonical_streamer_id') or ''),
            nickname=str(profile.get('nickname') or ''),
            total_income=0,
            voice_call_income=0,
            source='zero',
        )
        row['data_status'] = '已覆盖' if day_text in covered else '未覆盖'
        result.append(row)
        cursor -= timedelta(days=1)
    return result


def build_streamer_history_xlsx(
    *,
    profile: Dict[str, Any],
    date_to: str,
    rows: Sequence[Dict[str, Any]],
) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = '收益汇总'
    app_label = 'Timo' if profile.get('app_name') == 'timo' else 'Linky'
    guild_display_name = (
        timo_guild_display_name(profile.get('guild_name'))
        if profile.get('app_name') == 'timo'
        else str(profile.get('guild_name') or '')
    )
    covered_dates = [str(row['stat_date']) for row in rows if row.get('data_status') == '已覆盖']
    total_income = sum(_number(row.get('total_income')) for row in rows if row.get('data_status') == '已覆盖')
    one_to_one = sum(_number(row.get('one_to_one_income')) for row in rows if row.get('data_status') == '已覆盖')
    voice_call = sum(_number(row.get('voice_call_income')) for row in rows if row.get('data_status') == '已覆盖')
    summary_rows = [
        ('平台', app_label),
        ('目标公会', guild_display_name),
        ('主播 SID', str(profile.get('requested_streamer_id') or '')),
        ('主播昵称', str(profile.get('nickname') or '')),
        ('首次加入目标公会日期', str(profile.get('first_join_date') or '')),
        ('导出截止日期', date_to),
        ('实际数据覆盖开始', min(covered_dates) if covered_dates else ''),
        ('实际数据覆盖结束', max(covered_dates) if covered_dates else ''),
        ('未覆盖天数', sum(1 for row in rows if row.get('data_status') != '已覆盖')),
        ('总收益', total_income),
        ('1v1 收益（总收益-语音通话）', one_to_one),
        ('语音通话收益', voice_call),
        ('导出时间（北京时间）', datetime.now(timezone.utc).astimezone(BJ).strftime('%Y-%m-%d %H:%M:%S')),
    ]
    for row in summary_rows:
        summary.append(row)
    summary.column_dimensions['A'].width = 30
    summary.column_dimensions['B'].width = 34
    for cell in summary['A']:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='EAF1FF')
    for row_number in (10, 11, 12):
        summary.cell(row=row_number, column=2).number_format = '#,##0.00'
    summary.cell(row=3, column=2).number_format = '@'
    summary.cell(row=3, column=2).quotePrefix = True

    detail = workbook.create_sheet('每日收益')
    headers = ('日期', '主播 SID', '主播昵称', '目标公会', '总收益', '1v1 收益', '语音通话收益', '数据状态')
    detail.append(headers)
    for row in rows:
        detail.append((
            row.get('stat_date'),
            str(profile.get('requested_streamer_id') or ''),
            row.get('nickname') or profile.get('nickname') or '',
            guild_display_name,
            _number(row.get('total_income')),
            _number(row.get('one_to_one_income')),
            _number(row.get('voice_call_income')),
            row.get('data_status') or '',
        ))
    detail.freeze_panes = 'A2'
    detail.auto_filter.ref = detail.dimensions
    widths = (13, 20, 22, 24, 16, 16, 18, 12)
    for index, width in enumerate(widths, start=1):
        detail.column_dimensions[chr(64 + index)].width = width
    for cell in detail[1]:
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = PatternFill('solid', fgColor='1F4FA3')
        cell.alignment = Alignment(horizontal='center')
    for row_number in range(2, detail.max_row + 1):
        detail.cell(row=row_number, column=2).number_format = '@'
        detail.cell(row=row_number, column=2).quotePrefix = True
        for column in (5, 6, 7):
            detail.cell(row=row_number, column=column).number_format = '#,##0.00'

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
